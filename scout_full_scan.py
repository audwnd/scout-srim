# -*- coding: utf-8 -*-
"""
scout_full_scan.py  v2
======================
SCOUT 전종목 실수집 + S-RIM 계산 + 오류 자동 검토 스크립트

사용법:
  python scout_full_scan.py

결과:
  OUTPUT/scan_report_YYYYMMDD_HHMM.xlsx  ← 오류 리포트 (3개 시트)
  OUTPUT/scan_log_YYYYMMDD_HHMM.txt      ← 전체 로그

옵션 (파일 상단 CONFIG 수정):
  LIMIT      = 0      → 0이면 전종목, 숫자면 해당 개수만 테스트
  SLEEP_SEC  = 3      → 종목당 딜레이 (초). FnGuide 차단 방지용. 최소 2 권장
  SKIP_DONE  = True   → 이미 WORK 폴더에 JSON 있는 종목은 수집 건너뜀
  MARKET     = "ALL"  → "KOSPI" / "KOSDAQ" / "ALL"
"""

import sys, os, json, time, traceback, statistics
from pathlib import Path
from datetime import datetime

# ── 경로 설정 (app.py와 같은 폴더에 놓고 실행) ──────────────────
BASE = Path(__file__).parent

# ── CONFIG ──────────────────────────────────────────────────────
LIMIT     = 0        # 0 = 전종목. 테스트시 예: 50
SLEEP_SEC = 3        # 종목간 딜레이 (초). 2 이상 권장
SKIP_DONE = True     # WORK 폴더에 JSON 이미 있으면 스킵
MARKET    = "ALL"    # "KOSPI" / "KOSDAQ" / "ALL"
# ────────────────────────────────────────────────────────────────

sys.path.insert(0, str(BASE))

LOG_TIME  = datetime.now().strftime("%Y%m%d_%H%M")
OUT_DIR   = BASE / "OUTPUT"
WORK_DIR  = BASE / "WORK"
OUT_DIR.mkdir(exist_ok=True)
WORK_DIR.mkdir(exist_ok=True)

LOG_PATH  = OUT_DIR / f"scan_log_{LOG_TIME}.txt"
XLSX_PATH = OUT_DIR / f"scan_report_{LOG_TIME}.xlsx"


# ── 로거 ────────────────────────────────────────────────────────
class Logger:
    def __init__(self, path):
        self.f = open(str(path), "w", encoding="utf-8")

    def log(self, msg, also_print=True):
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {msg}"
        self.f.write(line + "\n")
        self.f.flush()
        if also_print:
            print(line)

    def close(self):
        self.f.close()


# ── ke 1회 수집 ─────────────────────────────────────────────────
def fetch_ke() -> float:
    """KIS BBB- 5년 수익률 — 전종목 공유 (1회만 수집)"""
    try:
        import kis_collector as _kis, importlib
        importlib.reload(_kis)
        ke = _kis.get_bbb_minus_5yr()
        ke = round(max(0.05, min(ke, 0.25)), 4)
        print(f"  [ke] BBB- 5년 수익률 수집 완료: {ke*100:.2f}%")
        return ke
    except Exception as e:
        print(f"  [ke] 수집 실패 → 기본값 10.31% 사용 ({e})")
        return 0.1031


# ── 종목 리스트 로드 ─────────────────────────────────────────────
def load_stocks():
    cache = WORK_DIR / "stock_list.json"
    if not cache.exists():
        print(f"[오류] {cache} 파일이 없습니다.")
        print("  → 서버를 한번 실행(python app.py)하면 자동 생성됩니다.")
        sys.exit(1)

    raw = json.loads(cache.read_text(encoding="utf-8"))
    stocks = []

    if isinstance(raw, dict) and "data" in raw:
        for name, val in raw["data"].items():
            if isinstance(val, (list, tuple)) and len(val) >= 1:
                code   = val[0]
                market = val[1] if len(val) > 1 else "KOSPI"
                stocks.append((name.strip(), code, market))
    elif isinstance(raw, list):
        for item in raw:
            if isinstance(item, dict):
                stocks.append((item.get("name",""), item.get("code",""), item.get("market","")))

    if MARKET != "ALL":
        stocks = [(n, c, m) for n, c, m in stocks if m == MARKET]
    if LIMIT > 0:
        stocks = stocks[:LIMIT]

    return stocks


# ── 빈 result 초기화 ─────────────────────────────────────────────
def _empty_result(name, code, market):
    return {
        # 기본 정보
        "종목명": name, "코드": code, "시장": market,
        "상태": "", "오류내용": "",
        # 수집 데이터
        "현재가": None, "베타": None, "발행주식수": None,
        # 연간 데이터
        "ROE_최근": None, "EPS_최근": None, "BPS_최근": None,
        "DPS_최근": None, "배당수익률": None,
        "영업이익_최근": None, "매출액_최근": None,
        # 컨센서스
        "컨센ROE": None, "컨센EPS": None,
        # ── S-RIM 계산 결과 ──────────────────────────────────
        "ke": None,
        "가중ROE": None,
        "적정주가": None,
        "매도가격": None,
        "매수가격": None,
        "현재가대비": None,   # (현재가/적정주가 - 1) × 100 (%)
        "배열": None,
        "ROE추세": None,
        # ── ROE 안정성 ───────────────────────────────────────
        "ROE표준편차": None,
        "ROE변동범위": None,
        "ROE안정성": None,
        # ── 이자보상배율 ─────────────────────────────────────
        "이자보상배율": None,
        "이자비용_최근": None,
        # ── FCF ─────────────────────────────────────────────
        "영업CF_최근": None,
        "CAPEX_최근": None,
        "FCF_최근": None,      # 억원
        "FCF수익률": None,     # %
        # ── 배당 ─────────────────────────────────────────────
        "시가배당수익률": None,
        "예상시가배당수익률": None,
        # ── 검증 플래그 ──────────────────────────────────────
        "ROE_단위오류":       False,
        "BPS_없음":           False,
        "EPS_없음":           False,
        "매출액_없음_비금융": False,
        "연도수_부족":        False,
        "컨센_없음":          False,
        "DPS_None인데_EPS있음": False,
        "영업이익률_극단":    False,
        "ROE_극단":           False,
        "PER_극단":           False,
        "시가총액_0":         False,
        "Finance_없음":       False,
        # ── 신규 플래그 ──────────────────────────────────────
        "적정주가_0":         False,   # S-RIM 계산불가 (ROE<ke 또는 데이터 부족)
        "적정주가_극단":      False,   # 현재가 대비 이상 괴리
        "이자보상배율_위험":  False,   # ICR < 1.5 (이자비용 있는 종목)
        "영업CF_없음":        False,   # FCF 계산 불가
        "경고목록": [],
    }


# ── 단일 종목 수집 + 검증 ────────────────────────────────────────
def scan_one(name, code, market, ke, logger):
    result = _empty_result(name, code, market)
    json_path = WORK_DIR / f"{code}_{name}.json"

    # ── 1) 수집 ──────────────────────────────────────────────
    if SKIP_DONE and json_path.exists():
        logger.log(f"  [SKIP] {name}({code})")
        try:
            data = json.loads(json_path.read_text(encoding="utf-8"))
        except Exception as e:
            result["상태"] = "JSON_읽기오류"
            result["오류내용"] = str(e)
            return result
    else:
        try:
            import fnguide_collector_v4 as _col, importlib
            importlib.reload(_col)
            data = _col.collect(name, code)
            m = data.get("meta", {}).get("market") or market
            data["_market"] = m
            with open(str(json_path), "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            try:
                import requests as _rq
                is_net = isinstance(e, (_rq.ConnectionError, _rq.Timeout))
            except Exception:
                is_net = False
            result["상태"] = "수집오류_네트워크" if is_net else "수집오류"
            result["오류내용"] = str(e)[:200]
            return result

    # ── 2) 데이터 추출 ───────────────────────────────────────
    ann = data.get("annual", {})
    con = data.get("consensus", {})
    fin = data.get("finance", {})

    def _last(lst):
        return next((v for v in reversed(lst or []) if v is not None), None)

    현재가     = data.get("현재가") or 0
    발행주식수 = data.get("발행주식수_보통") or 0
    베타       = data.get("베타")
    roe_list   = ann.get("ROE", [])
    eps_list   = ann.get("EPS", [])
    bps_list   = ann.get("BPS", [])
    dps_list   = ann.get("DPS", [])
    op_list    = ann.get("영업이익", [])
    rev_list   = ann.get("매출액", [])
    years      = ann.get("years", [])
    is_finance = not bool(rev_list)

    result["현재가"]        = 현재가
    result["베타"]          = 베타
    result["발행주식수"]    = 발행주식수
    result["ROE_최근"]      = round(_last(roe_list) * 100, 2) if _last(roe_list) is not None else None
    result["EPS_최근"]      = _last(eps_list)
    result["BPS_최근"]      = _last(bps_list)
    result["DPS_최근"]      = _last(dps_list)
    result["배당수익률"]    = data.get("배당수익률")
    result["영업이익_최근"] = _last(op_list)
    result["매출액_최근"]   = _last(rev_list)
    result["컨센ROE"]       = round(_last(con.get("ROE", [])) * 100, 2) if _last(con.get("ROE", [])) is not None else None
    result["컨센EPS"]       = _last(con.get("EPS", []))
    result["ke"]            = round(ke * 100, 2)

    warns = []

    # ── 3) 기본 검증 ─────────────────────────────────────────

    # ① 현재가 없음
    if not 현재가:
        result["상태"] = "현재가없음"
        result["오류내용"] = "현재가=0 (거래정지/상장폐지 의심)"
        return result

    # ② 시가총액 0
    시가총액 = round(현재가 * 발행주식수 / 1e8) if 발행주식수 else 0
    if 시가총액 == 0:
        result["시가총액_0"] = True
        warns.append("시가총액=0 (발행주식수 없음)")

    # ③ 연도 데이터 수 부족
    if len(years) < 3:
        result["연도수_부족"] = True
        warns.append(f"실적 연도 {len(years)}개 (3개 미만)")

    # ④ ROE 단위 오류
    roe_valid = [v for v in roe_list if v is not None]
    if roe_valid and max(abs(v) for v in roe_valid) > 2:
        result["ROE_단위오류"] = True
        warns.append(f"ROE 단위 이상: max={max(abs(v) for v in roe_valid):.2f}")

    # ⑤ ROE 극단값
    if roe_valid and max(abs(v) for v in roe_valid) > 1.0:
        result["ROE_극단"] = True
        warns.append(f"ROE 극단값: {[round(v*100,1) for v in roe_valid]}%")

    # ⑥ BPS 없음
    bps_valid = [v for v in bps_list if v is not None]
    if not bps_valid:
        result["BPS_없음"] = True
        warns.append("BPS 전체 None")

    # ⑦ EPS 없음
    eps_valid = [v for v in eps_list if v is not None]
    if not eps_valid:
        result["EPS_없음"] = True
        warns.append("EPS 전체 None")

    # ⑧ 비금융주 매출액 없음
    if not is_finance and not rev_list:
        result["매출액_없음_비금융"] = True
        warns.append("비금융주인데 매출액 없음")

    # ⑨ 컨센서스 없음
    if not con.get("years"):
        result["컨센_없음"] = True
        warns.append("컨센서스 없음")

    # ⑩ DPS 전체 None인데 EPS 있음
    dps_valid = [v for v in dps_list if v is not None]
    if eps_valid and not dps_valid and not is_finance:
        result["DPS_None인데_EPS있음"] = True
        warns.append("DPS 전체 None (배당 없는 종목이거나 수집 오류)")

    # ⑪ 영업이익률 극단 (비금융주)
    if not is_finance and op_list and rev_list:
        margins = [round(o/r*100, 1) for o, r in zip(op_list, rev_list)
                   if o is not None and r and r > 0]
        if margins and max(abs(m) for m in margins) > 80:
            result["영업이익률_극단"] = True
            warns.append(f"영업이익률 극단값: {margins}")

    # ⑫ PER 극단
    per = data.get("PER")
    if per and (per > 500 or per < 0):
        result["PER_극단"] = True
        warns.append(f"PER 극단값: {per}")

    # ⑬ Finance 데이터 없음
    if not fin or not fin.get("years"):
        result["Finance_없음"] = True
        warns.append("Finance 데이터 없음")

    # ── 4) S-RIM 계산 ────────────────────────────────────────
    try:
        import app as _app, importlib
        importlib.reload(_app)

        apt, sell, buy, meta = _app._srim_python(data, ke)
        result["적정주가"]  = apt
        result["매도가격"]  = sell
        result["매수가격"]  = buy
        result["가중ROE"]   = round(meta.get("roe", 0) * 100, 2) if meta.get("roe") is not None else None
        result["배열"]      = meta.get("배열", "")
        result["ROE추세"]   = meta.get("추세", "")

        if apt > 0 and 현재가 > 0:
            result["현재가대비"] = round((현재가 / apt - 1) * 100, 1)
            # ⑯ 적정주가 극단 판별 (현재가 대비 +300% 초과 or -80% 미만)
            괴리 = result["현재가대비"]
            if 괴리 < -80 or 괴리 > 300:
                result["적정주가_극단"] = True
                warns.append(f"적정주가 극단 괴리 {괴리:+.0f}% (적정={apt:,} / 현재가={현재가:,})")
        elif apt == 0:
            # ⑰ 적정주가 = 0
            result["적정주가_0"] = True
            warns.append("S-RIM 적정주가=0 (ROE<ke 또는 데이터 부족)")

    except Exception as e:
        warns.append(f"S-RIM 계산 오류: {e}")

    # ── 5) ROE 안정성 ────────────────────────────────────────
    try:
        roe_pct = [(v * 100 if abs(v) <= 2 else v) for v in roe_valid]
        if len(roe_pct) >= 2:
            std = round(statistics.stdev(roe_pct), 2)
            rng = round(max(roe_pct) - min(roe_pct), 2)
            stability = "안정" if std < 3 else ("보통" if std < 7 else "변동 큼")
            result["ROE표준편차"] = std
            result["ROE변동범위"] = rng
            result["ROE안정성"]   = stability
    except Exception:
        pass

    # ── 6) 이자보상배율 ──────────────────────────────────────
    try:
        interest_list = [v for v in (fin.get("이자비용") or []) if v is not None and v > 0]
        op_last       = _last(op_list)
        int_last      = interest_list[-1] if interest_list else None
        result["이자비용_최근"] = int_last

        if int_last and int_last > 0 and op_last is not None:
            icr = round(op_last / int_last, 2)
            result["이자보상배율"] = icr
            # ⑱ 이자보상배율 위험 (비금융주, 1.5배 미만)
            if not is_finance and icr < 1.5:
                result["이자보상배율_위험"] = True
                warns.append(f"이자보상배율 {icr:.1f}배 (1.5 미만, 이자 부담 위험)")
        elif not int_last and not is_finance:
            # 이자비용 데이터 없으면 경고 아닌 참고용 메모만 (금융주 제외)
            pass
    except Exception:
        pass

    # ── 7) FCF ───────────────────────────────────────────────
    try:
        op_cf_list  = [v for v in (fin.get("영업CF") or []) if v is not None]
        capex_list  = [v for v in (fin.get("CAPEX")  or []) if v is not None]
        shrx        = data.get("발행주식수_보통", 0) or 0

        if op_cf_list:
            result["영업CF_최근"] = round(op_cf_list[-1])
        else:
            result["영업CF_없음"] = True
            warns.append("영업CF 없음 (FCF 계산 불가)")

        if capex_list:
            result["CAPEX_최근"] = round(capex_list[-1])

        if op_cf_list and capex_list and shrx > 0:
            n = min(len(op_cf_list), len(capex_list), 3)
            fcf_list = [o - c for o, c in zip(op_cf_list[-n:], capex_list[-n:])]
            w = list(range(1, n + 1))
            fcf_avg = sum(f * wt for f, wt in zip(fcf_list, w)) / sum(w)
            result["FCF_최근"] = round(fcf_avg)
            if 현재가 > 0 and shrx > 0:
                fcf_ps = fcf_avg * 1e8 / shrx
                result["FCF수익률"] = round(fcf_ps / 현재가 * 100, 2)
    except Exception:
        pass

    # ── 8) 시가배당수익률 ────────────────────────────────────
    try:
        dps_last_valid = _last(dps_list)
        con_dps1       = next((v for v in con.get("DPS", []) if v), None)
        if dps_last_valid and 현재가 > 0:
            result["시가배당수익률"]   = round(dps_last_valid / 현재가 * 100, 2)
        if con_dps1 and 현재가 > 0:
            result["예상시가배당수익률"] = round(con_dps1 / 현재가 * 100, 2)
    except Exception:
        pass

    # ── 9) 최종 상태 결정 ────────────────────────────────────
    critical_flags = [
        result["ROE_단위오류"],
        result["BPS_없음"],
        result["시가총액_0"],
        result["연도수_부족"],
        result["이자보상배율_위험"],
    ]
    warning_flags = [
        result["ROE_극단"],
        result["EPS_없음"],
        result["컨센_없음"],
        result["영업이익률_극단"],
        result["PER_극단"],
        result["적정주가_0"],
        result["적정주가_극단"],
        result["영업CF_없음"],
    ]

    if any(critical_flags):
        result["상태"] = "오류"
    elif any(warning_flags) or warns:
        result["상태"] = "경고"
    else:
        result["상태"] = "정상"

    result["경고목록"] = warns
    result["오류내용"] = " | ".join(warns) if warns else ""
    return result


# ── 엑셀 리포트 저장 ─────────────────────────────────────────────
def save_report(rows, path):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        print("[경고] openpyxl 없음 - pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill  = PatternFill("solid", fgColor="1F4E79")
    header_font  = Font(color="FFFFFF", bold=True, size=9)
    section_fill = PatternFill("solid", fgColor="2E75B6")
    section_font = Font(color="FFFFFF", bold=True, size=9)

    fill_ok   = PatternFill("solid", fgColor="E8F5E9")
    fill_warn = PatternFill("solid", fgColor="FFF9C4")
    fill_err  = PatternFill("solid", fgColor="FFEBEE")
    fill_skip = PatternFill("solid", fgColor="F5F5F5")

    # ── 컬럼 정의 ─────────────────────────────────────────────
    COLS = [
        # (헤더, key, 너비)
        ("번호",          "__idx__",              5),
        ("종목명",        "종목명",               14),
        ("코드",          "코드",                  8),
        ("시장",          "시장",                  7),
        ("상태",          "상태",                  7),
        # 현재가/기본
        ("현재가",        "현재가",                9),
        ("시가총액(억)",   "__mktcap__",           10),
        ("베타",          "베타",                  7),
        # S-RIM
        ("ke(%)",         "ke",                    7),
        ("가중ROE(%)",    "가중ROE",               9),
        ("적정주가",      "적정주가",              10),
        ("매도가격",      "매도가격",              10),
        ("매수가격",      "매수가격",              10),
        ("현재가대비(%)", "현재가대비",             9),
        ("배열",          "배열",                  7),
        ("ROE추세",       "ROE추세",               8),
        # ROE 안정성
        ("ROE최근(%)",    "ROE_최근",               9),
        ("ROE표준편차",   "ROE표준편차",            9),
        ("ROE변동범위",   "ROE변동범위",            9),
        ("ROE안정성",     "ROE안정성",              9),
        # 이자보상배율
        ("이자보상배율",  "이자보상배율",          10),
        ("이자비용(억)",  "이자비용_최근",          9),
        # FCF
        ("영업CF(억)",    "영업CF_최근",            9),
        ("CAPEX(억)",     "CAPEX_최근",             9),
        ("FCF(억)",       "FCF_최근",               9),
        ("FCF수익률(%)",  "FCF수익률",              9),
        # 배당
        ("DPS",           "DPS_최근",               9),
        ("시가배당(%)",   "시가배당수익률",          9),
        ("예상시가배당(%)","예상시가배당수익률",    11),
        ("배당수익률(%)", "배당수익률",              9),
        # 기타 재무
        ("EPS",           "EPS_최근",               9),
        ("BPS",           "BPS_최근",               9),
        ("영업이익(억)",  "영업이익_최근",          10),
        ("매출액(억)",    "매출액_최근",            10),
        ("컨센ROE(%)",    "컨센ROE",                 9),
        ("컨센EPS",       "컨센EPS",                 9),
        # 검증 플래그
        ("ROE단위오류",   "ROE_단위오류",            9),
        ("ROE극단",       "ROE_극단",                7),
        ("BPS없음",       "BPS_없음",                7),
        ("EPS없음",       "EPS_없음",                7),
        ("연도수부족",    "연도수_부족",             8),
        ("컨센없음",      "컨센_없음",               7),
        ("적정주가=0",    "적정주가_0",              9),
        ("적정주가극단",  "적정주가_극단",            9),
        ("이자보상위험",  "이자보상배율_위험",        9),
        ("영업CF없음",    "영업CF_없음",              8),
        ("Finance없음",   "Finance_없음",             9),
        ("오류내용",      "오류내용",               45),
    ]

    def _write_header(ws):
        for ci, (h, key, w) in enumerate(COLS, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = header_fill; c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
            ws.column_dimensions[c.column_letter].width = w
        ws.row_dimensions[1].height = 32
        ws.freeze_panes = "F2"

    def _row_values(row, idx):
        vals = []
        현재가 = row.get("현재가") or 0
        발행주식수 = row.get("발행주식수") or 0
        시가총액 = round(현재가 * 발행주식수 / 1e8) if (현재가 and 발행주식수) else None
        for h, key, w in COLS:
            if key == "__idx__":
                vals.append(idx)
            elif key == "__mktcap__":
                vals.append(시가총액)
            elif key in ("ROE_단위오류","ROE_극단","BPS_없음","EPS_없음","연도수_부족",
                         "컨센_없음","적정주가_0","적정주가_극단","이자보상배율_위험",
                         "영업CF_없음","Finance_없음"):
                vals.append("Y" if row.get(key) else "")
            else:
                vals.append(row.get(key))
        return vals

    def _write_row(ws, ri, row, idx):
        상태 = row.get("상태", "")
        if 상태 == "정상":     fill = fill_ok
        elif 상태 == "경고":   fill = fill_warn
        elif "오류" in 상태:   fill = fill_err
        else:                  fill = fill_skip

        for ci, val in enumerate(_row_values(row, idx), 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill; c.border = border
            c.font = Font(size=9)
            c.alignment = Alignment(vertical="center",
                                    wrap_text=(ci == len(COLS)))
            # 수치 포맷
            if isinstance(val, float):
                c.number_format = "#,##0.00"
            elif isinstance(val, int) and ci > 5:
                c.number_format = "#,##0"

    # ── 시트1: 전체 결과 ──────────────────────────────────────
    ws = wb.active
    ws.title = "전체결과"
    _write_header(ws)
    for ri, row in enumerate(rows, 2):
        _write_row(ws, ri, row, ri - 1)

    # ── 시트2: 오류/경고 종목만 ──────────────────────────────
    ws2 = wb.create_sheet("오류경고")
    _write_header(ws2)
    ri2 = 2
    err_states = {"오류","경고","수집오류","수집오류_네트워크","JSON_읽기오류","현재가없음"}
    for row in rows:
        if row.get("상태") in err_states:
            _write_row(ws2, ri2, row, ri2 - 1)
            ri2 += 1

    # ── 시트3: S-RIM 유효 종목 (적정주가 > 0, 현재가대비 포함) ──
    ws3 = wb.create_sheet("S-RIM결과")
    _write_header(ws3)
    srim_rows = sorted(
        [r for r in rows if r.get("적정주가") and r.get("적정주가") > 0],
        key=lambda x: (x.get("현재가대비") or 999)
    )
    for ri3, row in enumerate(srim_rows, 2):
        _write_row(ws3, ri3, row, ri3 - 1)

    # ── 시트4: 요약 ──────────────────────────────────────────
    ws4 = wb.create_sheet("요약")
    total  = len(rows)
    ok     = sum(1 for r in rows if r.get("상태") == "정상")
    warn   = sum(1 for r in rows if r.get("상태") == "경고")
    err    = sum(1 for r in rows if "오류" in r.get("상태",""))
    other  = total - ok - warn - err

    srim_valid = sum(1 for r in rows if r.get("적정주가") and r.get("적정주가") > 0)
    undervalued = sum(1 for r in rows if (r.get("현재가대비") or 0) < -20)  # 적정가 대비 20% 이하
    overvalued  = sum(1 for r in rows if (r.get("현재가대비") or 0) > 20)

    summary = [
        ["항목",                "건수",  "비율 / 비고"],
        ["전체 종목",           total,   "100%"],
        ["✅ 정상",              ok,      f"{ok/total*100:.1f}%" if total else "0%"],
        ["⚠️ 경고",             warn,    f"{warn/total*100:.1f}%" if total else "0%"],
        ["❌ 오류",              err,     f"{err/total*100:.1f}%" if total else "0%"],
        ["기타(수집불가 등)",   other,   f"{other/total*100:.1f}%" if total else "0%"],
        [],
        ["── S-RIM 결과 ──", "", ""],
        ["적정주가 산출 성공",  srim_valid,  f"{srim_valid/total*100:.1f}%" if total else ""],
        ["저평가 (-20% 이하)",  undervalued, f"적정주가 대비 현재가 낮음"],
        ["고평가 (+20% 이상)",  overvalued,  f"적정주가 대비 현재가 높음"],
        [],
        ["── 오류 유형별 ──", "", ""],
        ["ROE 단위오류",        sum(1 for r in rows if r.get("ROE_단위오류")),    ""],
        ["BPS 없음",            sum(1 for r in rows if r.get("BPS_없음")),        ""],
        ["EPS 없음",            sum(1 for r in rows if r.get("EPS_없음")),        ""],
        ["연도수 부족",         sum(1 for r in rows if r.get("연도수_부족")),     ""],
        ["컨센서스 없음",       sum(1 for r in rows if r.get("컨센_없음")),       ""],
        ["적정주가 = 0",        sum(1 for r in rows if r.get("적정주가_0")),      "ROE < ke"],
        ["적정주가 극단 괴리",  sum(1 for r in rows if r.get("적정주가_극단")),   "현재가 대비 ±극단"],
        ["이자보상배율 위험",   sum(1 for r in rows if r.get("이자보상배율_위험")),"ICR < 1.5배"],
        ["영업CF 없음",         sum(1 for r in rows if r.get("영업CF_없음")),     "FCF 계산 불가"],
        ["Finance 없음",        sum(1 for r in rows if r.get("Finance_없음")),    ""],
        ["ROE 극단값",          sum(1 for r in rows if r.get("ROE_극단")),        ""],
        ["PER 극단값",          sum(1 for r in rows if r.get("PER_극단")),        ""],
    ]

    for ri, row_data in enumerate(summary, 1):
        if not row_data:
            continue
        for ci, val in enumerate(row_data, 1):
            c = ws4.cell(row=ri, column=ci, value=val)
            if ri == 1 or "──" in str(row_data[0]):
                c.fill = header_fill if ri == 1 else section_fill
                c.font = header_font if ri == 1 else section_font
            if isinstance(val, int):
                c.number_format = "#,##0"
            c.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                    vertical="center")

    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 10
    ws4.column_dimensions["C"].width = 24

    wb.save(str(path))
    print(f"\n  ✓ 리포트 저장: {path}")


# ── 메인 ────────────────────────────────────────────────────────
def main():
    logger = Logger(LOG_PATH)
    logger.log("=" * 60)
    logger.log("  SCOUT 전종목 검토 v2 시작")
    logger.log(f"  설정: LIMIT={LIMIT}, SLEEP={SLEEP_SEC}초, SKIP={SKIP_DONE}, MARKET={MARKET}")
    logger.log("=" * 60)

    # ke 1회 수집
    ke = fetch_ke()
    logger.log(f"  요구수익률(ke): {ke*100:.2f}%")

    stocks = load_stocks()
    total  = len(stocks)
    logger.log(f"  대상 종목: {total}개\n")

    results  = []
    ok_cnt   = warn_cnt = err_cnt = 0

    for i, (name, code, market) in enumerate(stocks, 1):
        logger.log(f"[{i:>4}/{total}] {name} ({code}) {market}")

        try:
            row = scan_one(name, code, market, ke, logger)
        except Exception as e:
            row = _empty_result(name, code, market)
            row["상태"] = "수집오류"
            row["오류내용"] = str(e)[:200]
            logger.log(f"  [예외] {e}")

        상태 = row.get("상태", "")
        if 상태 == "정상":     ok_cnt += 1;   icon = "✅"
        elif 상태 == "경고":   warn_cnt += 1; icon = "⚠️"
        else:                  err_cnt += 1;  icon = "❌"

        apt  = row.get("적정주가")
        roe  = row.get("가중ROE")
        icr  = row.get("이자보상배율")
        dv   = row.get("현재가대비")

        logger.log(
            f"  {icon} {상태:<6} | "
            f"ROE={roe}% | 적정={apt:,} | 대비={dv:+.0f}%"
            if apt and dv is not None else
            f"  {icon} {상태:<6} | ROE={roe}% | 적정주가 산출불가"
        )
        if row.get("오류내용"):
            logger.log(f"     → {row['오류내용'][:120]}")

        results.append(row)

        # 100개마다 중간 저장
        if i % 100 == 0:
            logger.log(f"\n  ── 중간 저장 ({i}/{total}) ──")
            save_report(results, XLSX_PATH)

        if i < total:
            time.sleep(SLEEP_SEC)

    # ── 최종 리포트 ──────────────────────────────────────────
    logger.log("\n" + "=" * 60)
    logger.log(f"  완료: 전체 {total} | 정상 {ok_cnt} | 경고 {warn_cnt} | 오류 {err_cnt}")
    logger.log("=" * 60)

    save_report(results, XLSX_PATH)
    logger.log(f"  리포트: {XLSX_PATH}")
    logger.log(f"  로그:   {LOG_PATH}")
    logger.close()

    print(f"\n  리포트: {XLSX_PATH}")


if __name__ == "__main__":
    main()
