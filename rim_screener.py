# -*- coding: utf-8 -*-
"""
rim_screener.py - RIM병합 프로젝트 스크리닝 엔진 v2
필터 체계:
  [pykrx]    시가총액 KOSPI 500억↑ / KOSDAQ 300억↑, 주가 1000원↑, 거래정지 제외
  [XML 빠름] 해외기업 제외, 영업이익 2년↑ 손실 제외, 법인세차감전 1년↑ 손실 제외
  [KRX]      소수계좌매도 공시 종목 제외
  [Finance]  단기차입금 > 이익잉여금 제외 (2단계에서 적용)
"""

import json, sys, importlib, threading, re, subprocess
from pathlib import Path
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from bs4 import BeautifulSoup


# ══════════════════════════════════════════════
# pykrx 자동 버전 체크 & 업그레이드
# ══════════════════════════════════════════════

def _ensure_pykrx_latest() -> None:
    """
    pykrx 최신 버전 확인 → 구버전이면 자동 업그레이드.
    스크리닝 시작 전 1회 실행. 업그레이드 후 모듈 재로드.
    """
    try:
        import pykrx
        import importlib.metadata as _meta
        current = _meta.version("pykrx")

        # PyPI 최신 버전 조회
        r = requests.get("https://pypi.org/pypi/pykrx/json", timeout=5)
        latest = r.json()["info"]["version"]

        if current == latest:
            print(f"  [pykrx] v{current} (최신)")
            return

        print(f"  [pykrx] 업그레이드: v{current} → v{latest} ...")
        subprocess.run(
            [sys.executable, "-m", "pip", "install", "--upgrade", "pykrx", "-q"],
            check=True, capture_output=True
        )
        # 업그레이드된 버전 재로드
        importlib.reload(pykrx)
        import importlib.metadata as _meta2
        new_ver = _meta2.version("pykrx")
        print(f"  [pykrx] 업그레이드 완료 → v{new_ver}")

    except Exception as e:
        print(f"  [pykrx] 버전 체크 실패 (무시): {e}")

BASE      = Path(__file__).parent
CACHE_DIR = BASE / "WORK" / "xml_cache"
CACHE_DIR.mkdir(parents=True, exist_ok=True)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": "https://comp.fnguide.com"
}
MAX_WORKERS    = 10
XML_CACHE_DAYS = 7


# ══════════════════════════════════════════════
# 1. 할인율
# ══════════════════════════════════════════════

def get_discount_rate() -> float:
    cache_f = CACHE_DIR / "ke.json"
    today   = datetime.now().strftime("%Y-%m-%d")
    try:
        if cache_f.exists():
            c = json.loads(cache_f.read_text())
            if c.get("date") == today:
                return c["rate"]
    except Exception:
        pass
    try:
        from io import BytesIO
        import pandas as pd
        r  = requests.post(
            "https://www.kisrating.com/ratingsStatistics/statics_spread.do",
            data={}, headers=HEADERS, timeout=10)
        df = pd.read_html(BytesIO(r.content), header=0)[0].set_index("구분")
        ke = float(df.loc["BBB-", "5년"]) / 100
        cache_f.write_text(json.dumps({"date": today, "rate": ke}))
        print(f"  할인율(BBB- 5년): {ke*100:.2f}%")
        return ke
    except Exception:
        print(f"  [할인율 조회 실패] → 기본값 10.26%")
        return 0.1026


# ══════════════════════════════════════════════
# 2. 소수계좌매도 공시 종목 (KRX KIND)
# ══════════════════════════════════════════════

def get_minority_sell_codes() -> set:
    """KRX KIND에서 소수계좌매도 공시 종목 코드 조회 (캐시: 1일)"""
    cache_f = CACHE_DIR / "minority_sell.json"
    today   = datetime.now().strftime("%Y-%m-%d")
    try:
        if cache_f.exists():
            c = json.loads(cache_f.read_text())
            if c.get("date") == today:
                return set(c.get("codes", []))
    except Exception:
        pass

    codes = set()
    try:
        # KRX KIND 소수계좌매도 공시 조회
        url  = "https://kind.krx.co.kr/disclosure/searchtodaydisclosure.do"
        data = {
            "method": "searchTodayDisclosureSub",
            "currentPageSize": "100",
            "pageIndex": "1",
            "orderMode": "0",
            "orderStat": "D",
            "forward": "todaydisclosure_sub",
            "disclosureType": "",
            "searchCodeType": "",
            "searchCorpName": "",
            "marketType": "",
            "reportNm": "소수계좌매도",
        }
        r    = requests.post(url, data=data, headers=HEADERS, timeout=10)
        soup = BeautifulSoup(r.text, "html.parser")
        for row in soup.select("table tr"):
            tds = row.find_all("td")
            if not tds: continue
            # 종목코드 추출
            for td in tds:
                link = td.find("a", href=True)
                if link and "gicode" in link.get("href",""):
                    code = link["href"].split("gicode=A")[-1][:6]
                    if code.isdigit():
                        codes.add(code)

        cache_f.write_text(json.dumps({"date": today, "codes": list(codes)}))
        print(f"  소수계좌매도 공시: {len(codes)}개")
    except Exception as e:
        print(f"  [소수계좌매도 조회 실패] {e}")

    return codes


# ══════════════════════════════════════════════
# 2-B. 관리종목·투자경고·투자위험 종목 조회
# ══════════════════════════════════════════════

def get_admin_codes() -> set:
    """
    KRX KIND 검색으로 관리종목·투자경고·투자위험·상장폐지사유 종목 코드 수집
    캐시: 1일 (각 키워드 최근 100건 기준 → 장기 지정종목은 일부 누락 가능)
    """
    cache_f = CACHE_DIR / "admin_codes.json"
    today   = datetime.now().strftime("%Y-%m-%d")
    try:
        if cache_f.exists():
            c = json.loads(cache_f.read_text())
            if c.get("date") == today:
                return set(c.get("codes", []))
    except Exception:
        pass

    codes = set()
    url   = "https://kind.krx.co.kr/disclosure/searchtodaydisclosure.do"
    ADMIN_KEYWORDS = [
        "관리종목지정",
        "투자경고종목지정",
        "투자위험종목지정",
        "상장폐지사유발생",
        "거래정지예고",
    ]

    for kw in ADMIN_KEYWORDS:
        try:
            r    = requests.post(url, headers=HEADERS, data={
                "method":          "searchTodayDisclosureSub",
                "currentPageSize": "100",
                "pageIndex":       "1",
                "orderMode":       "0",
                "orderStat":       "D",
                "forward":         "todaydisclosure_sub",
                "disclosureType":  "",
                "searchCodeType":  "",
                "searchCorpName":  "",
                "marketType":      "",
                "reportNm":        kw,
            }, timeout=10)
            soup = BeautifulSoup(r.text, "html.parser")
            for row in soup.select("table tr"):
                for td in row.find_all("td"):
                    a = td.find("a", href=True)
                    if a and "gicode=A" in a.get("href", ""):
                        code = a["href"].split("gicode=A")[-1][:6]
                        if code.isdigit():
                            codes.add(code)
        except Exception as e:
            print(f"  [관리종목 조회 오류] {kw}: {e}")

    try:
        cache_f.write_text(json.dumps({"date": today, "codes": list(codes)}))
    except Exception:
        pass
    print(f"  관리종목/투자경고/위험: {len(codes)}개")
    return codes


# ══════════════════════════════════════════════
# 2-C. 치명적 부정 공시 종목 사전 배제
# ══════════════════════════════════════════════

def get_critical_negative_codes() -> set:
    """
    횡령·배임·불성실공시·감사의견거절·한정 공시 종목 코드 조회
    → get_filtered_tickers()에서 완전 제외 (예외 없음)
    캐시: 1일 / KIND 최근 100건 기준
    """
    cache_f = CACHE_DIR / "critical_neg.json"
    today   = datetime.now().strftime("%Y-%m-%d")
    try:
        if cache_f.exists():
            c = json.loads(cache_f.read_text())
            if c.get("date") == today:
                return set(c.get("codes", []))
    except Exception:
        pass

    codes = set()
    url   = "https://kind.krx.co.kr/disclosure/searchtodaydisclosure.do"
    CRITICAL_KEYWORDS = [
        "횡령",
        "배임",
        "불성실공시",
        "감사보고서(의견거절)",
        "감사보고서(한정)",
        "감사의견거절",
        "한정의견",
    ]

    for kw in CRITICAL_KEYWORDS:
        try:
            r    = requests.post(url, headers=HEADERS, data={
                "method":          "searchTodayDisclosureSub",
                "currentPageSize": "100",
                "pageIndex":       "1",
                "orderMode":       "0",
                "orderStat":       "D",
                "forward":         "todaydisclosure_sub",
                "disclosureType":  "",
                "searchCodeType":  "",
                "searchCorpName":  "",
                "marketType":      "",
                "reportNm":        kw,
            }, timeout=10)
            soup = BeautifulSoup(r.text, "html.parser")
            for row in soup.select("table tr"):
                for td in row.find_all("td"):
                    a = td.find("a", href=True)
                    if a and "gicode=A" in a.get("href", ""):
                        code = a["href"].split("gicode=A")[-1][:6]
                        if code.isdigit():
                            codes.add(code)
        except Exception as e:
            print(f"  [치명적공시 조회 오류] {kw}: {e}")

    try:
        cache_f.write_text(json.dumps({"date": today, "codes": list(codes)}))
    except Exception:
        pass
    print(f"  횡령·배임·불성실·감사거절 종목: {len(codes)}개")
    return codes


# ══════════════════════════════════════════════
# 3-A. 수급강도 캐시 (배치 공유용)
# ══════════════════════════════════════════════

_SUPPLY_CACHE: dict = {}   # {code: result_dict} — 스크리닝 1회 동안 재사용

def _biz_daterange(n_days: int) -> tuple:
    """최근 n_days 영업일 (strt, end) 반환 (YYYYMMDD)"""
    dates = []
    d = datetime.now() - timedelta(days=1)
    while len(dates) < n_days + 5:   # 여유분 확보
        if d.weekday() < 5:
            dates.append(d.strftime("%Y%m%d"))
        d -= timedelta(days=1)
    dates.sort()
    return dates[-n_days] if len(dates) >= n_days else dates[0], dates[-1]


# ── 시장 전체 수급 데이터를 한 번에 수집해 캐시에 저장
_MARKET_SUPPLY_LOADED = False

def _load_market_supply_batch():
    """
    pykrx로 KOSPI+KOSDAQ 전종목 외기관 5일 순매수를 한 번에 수집.
    개별 호출 대신 전체 수집 후 캐시 → 속도 대폭 개선.
    """
    global _MARKET_SUPPLY_LOADED
    if _MARKET_SUPPLY_LOADED:
        return

    strt5, end = _biz_daterange(5)
    strt25, _  = _biz_daterange(25)

    try:
        from pykrx import stock as _s

        # ── 투자자별 순매수 (5일 합산) — 거래대금 우선, 거래량 폴백
        inv_map = {}  # {code: net_buy_억}

        def _get_inv_map(market: str) -> dict:
            """
            시장 전체 외국인·연기금·기관합계 5일 순매수를 한 번에 수집.
            get_market_net_purchases_of_equities 사용 — 인덱스=티커, 컬럼에 순매수거래대금 포함.
            반환: {code: {"foreign_buy": 억, "pension_buy": 억, "inst_buy": 억}}
            """
            result_map = {}
            investor_keys = [
                ("외국인",  "foreign_buy"),
                ("연기금",  "pension_buy"),
                ("기관합계","inst_buy"),
            ]
            for investor, key in investor_keys:
                try:
                    df = _s.get_market_net_purchases_of_equities(strt5, end, market, investor)
                    if df is None or df.empty:
                        continue
                    # 순매수거래대금 컬럼 찾기
                    col = next((c for c in df.columns if "순매수거래대금" in c), None)
                    if not col:
                        # 폴백: 순매수거래량 × 현재가는 어렵므로 순매수 관련 컬럼 시도
                        col = next((c for c in df.columns if "순매수" in c and "대금" in c), None)
                    if not col:
                        print(f"  [수급배치] {market}/{investor} 순매수거래대금 컬럼 없음: {list(df.columns)}")
                        continue
                    for idx in df.index:
                        c = str(idx).zfill(6)
                        result_map.setdefault(c, {})[key] = round(
                            float(df.loc[idx, col] or 0) / 1e8, 1)
                    print(f"  [수급배치] {market}/{investor} {len(df)}개 수집")
                except Exception as e:
                    print(f"  [수급배치] {market}/{investor} 오류: {e}")
            return result_map

        for market in ["KOSPI", "KOSDAQ"]:
            try:
                m_map = _get_inv_map(market)
                if m_map:
                    inv_map.update(m_map)
                    print(f"  [수급배치] {market} {len(m_map)}개 외국인·연기금·기관 수집 완료")
                else:
                    print(f"  [수급배치] {market} 데이터 없음 → 개별 호출로 대체")
            except Exception as e:
                print(f"  [수급배치] {market} 오류: {e}")

        if inv_map:
            for code, vals in inv_map.items():
                entry = _SUPPLY_CACHE.get(code, {})
                entry.update({
                    "foreign_buy": vals.get("foreign_buy"),
                    "pension_buy": vals.get("pension_buy"),
                    "inst_buy":    vals.get("inst_buy"),
                    "net_buy":     round(
                        (vals.get("foreign_buy") or 0) + (vals.get("inst_buy") or 0), 1
                    ),
                })
                _SUPPLY_CACHE[code] = entry

        # ── 연속 순매수일수 배치 계산 (일별 외국인+기관합계, 최근 5영업일)
        # 배치로 미리 계산해 두면 get_supply_strength()에서 개별 호출이 불필요 → 스레드 폭증 방지
        print("  [수급배치] 연속 순매수일수 계산 중...")
        # 최근 5 영업일 목록 구성 (최신→오래된)
        _biz5: list = []
        _d_tmp = datetime.strptime(end, "%Y%m%d")
        _strt5_dt = datetime.strptime(strt5, "%Y%m%d")
        while _d_tmp >= _strt5_dt and len(_biz5) < 5:
            if _d_tmp.weekday() < 5:
                _biz5.append(_d_tmp.strftime("%Y%m%d"))
            _d_tmp -= timedelta(days=1)
        # _biz5: 최신→오래된 순

        _daily_f: dict = {}   # {code: {date: float}}  외국인 일별 순매수
        _daily_i: dict = {}   # {code: {date: float}}  기관합계 일별 순매수

        for _day in reversed(_biz5):   # 오래된→최신 순으로 수집 (덮어쓰기 방지)
            for _mkt in ["KOSPI", "KOSDAQ"]:
                for _inv, _store in [("외국인", _daily_f), ("기관합계", _daily_i)]:
                    try:
                        _df_d = _s.get_market_net_purchases_of_equities(
                            _day, _day, _mkt, _inv)
                        if _df_d is None or _df_d.empty:
                            continue
                        _col_d = next(
                            (c for c in _df_d.columns if "순매수거래대금" in c), None)
                        if not _col_d:
                            continue
                        for _idx in _df_d.index:
                            _cc = str(_idx).zfill(6)
                            _store.setdefault(_cc, {})[_day] = float(
                                _df_d.loc[_idx, _col_d] or 0)
                    except Exception:
                        pass

        _all_daily = set(_daily_f.keys()) | set(_daily_i.keys())
        for _cc in _all_daily:
            _consec = 0
            for _day in _biz5:   # 최신→오래된 순
                _fv = _daily_f.get(_cc, {}).get(_day, 0)
                _iv = _daily_i.get(_cc, {}).get(_day, 0)
                if _fv > 0 or _iv > 0:
                    _consec += 1
                else:
                    break
            _entry = _SUPPLY_CACHE.get(_cc, {})
            _entry["consec_days"] = _consec
            _SUPPLY_CACHE[_cc] = _entry
        print(f"  [수급배치] 연속일수 {len(_all_daily)}개 종목 계산 완료")

        # ── 공매도 잔고비율 배치 수집
        today = _prev_business_day()
        for market in ["KOSPI", "KOSDAQ"]:
            try:
                df_short = _s.get_shorting_balance_by_ticker(today, market=market)
                if df_short is None or df_short.empty:
                    continue
                ratio_col = next((c for c in df_short.columns if "비중" in c), None)
                if not ratio_col:
                    continue
                for idx in df_short.index:
                    c = str(idx).zfill(6)
                    entry = _SUPPLY_CACHE.get(c, {})
                    entry["short_ratio"] = round(float(df_short.loc[idx, ratio_col] or 0), 2)
                    _SUPPLY_CACHE[c] = entry
                print(f"  [공매도] {market} {len(df_short)}개 잔고비율 수집")
            except Exception as e:
                print(f"  [공매도] {market} 오류: {e}")

    except Exception as e:
        print(f"  [수급 배치 수집 오류] {e}")

    _MARKET_SUPPLY_LOADED = True


# ══════════════════════════════════════════════
# 3-B. 수급강도 개별 종목 조회
# ══════════════════════════════════════════════

def get_supply_strength(code: str, price: float) -> dict:
    """
    수급강도 체크 v2 — 5-factor 분석

    수집 항목:
    ① 외국인 5일 순매수(억)
    ② 연기금 5일 순매수(억)  ← 핵심 장기 신호
    ③ 기관합계 5일 순매수(억)
    ④ 연속 순매수일수 (외국인 OR 연기금 기준, 둘 다 매도면 초기화)
    ⑤ 거래량비율 (최근 5일 / 20일 평균)
    ⑥ 가격 방향성 (최근 5일 중 상승일 수)

    수급 등급:
    ★★★: 외국인+연기금 동반순매수 AND 연속 3일↑ AND 거래량 1.5배↑ AND 상승 3일↑
    ★★ : 외국인+기관합계 동반순매수 AND 연속 2일↑ AND 상승 3일↑
    ★  : 외국인 OR 연기금 한쪽 순매수 AND 상승 3일↑

    is_strong = ★★ 이상 (예외B 발동 기준)
    """
    if code in _SUPPLY_CACHE and _SUPPLY_CACHE[code].get("_done"):
        return _SUPPLY_CACHE[code]

    # 배치 캐시에서 부분 데이터 인계
    cached = _SUPPLY_CACHE.get(code, {})
    result = {
        "foreign_buy":   cached.get("foreign_buy"),    # 외국인 5일 순매수(억)
        "pension_buy":   cached.get("pension_buy"),    # 연기금 5일 순매수(억)
        "inst_buy":      cached.get("inst_buy"),       # 기관합계 5일 순매수(억)
        "net_buy":       cached.get("net_buy"),        # 외국인+기관합계(기존 호환)
        "vol_ratio":     cached.get("vol_ratio"),      # 거래량비율
        "consec_days":   cached.get("consec_days", 0), # 연속 순매수일수
        "price_up_days": cached.get("price_up_days", 0), # 5일 중 상승일수
        "short_ratio":   cached.get("short_ratio"),    # 공매도 잔고비중(%)
        "grade":         "",
        "is_strong":     False,
        "label":         "",
    }

    strt5,  _    = _biz_daterange(5)
    strt10, end  = _biz_daterange(10)
    strt25, _    = _biz_daterange(25)

    # ① 투자자별 순매수 + 연속일수 (배치 캐시 미수집분만 호출)
    # 배치에서 foreign_buy가 채워졌으면 consec_days도 배치에서 수집됨 → 개별 호출 불필요
    # → need_investor = True인 경우만 개별 pykrx 호출 (동시 요청 폭증 방지)
    need_investor = result["foreign_buy"] is None
    if need_investor:
        # pykrx 개별 호출은 KRX 서버 응답 없을 때 무한 대기 → 10초 타임아웃 스레드로 감쌈
        import threading as _th

        def _do_investor_fetch():
            try:
                from pykrx import stock as _s

                # 방법A: 5일 합산 투자자별 순매수
                try:
                    df_sum = _s.get_market_trading_value_by_investor(strt5, end, code)
                    if df_sum is not None and not df_sum.empty:
                        순매수_col = next((c for c in df_sum.columns if "순매수" in c), None)
                        if 순매수_col:
                            if result["foreign_buy"] is None:
                                for inv in ["외국인합계", "외국인"]:
                                    if inv in df_sum.index:
                                        result["foreign_buy"] = round(
                                            float(df_sum.loc[inv, 순매수_col] or 0) / 1e8, 1)
                                        break
                            if result["pension_buy"] is None:
                                for inv in ["연기금 등", "연기금등", "연기금"]:
                                    if inv in df_sum.index:
                                        result["pension_buy"] = round(
                                            float(df_sum.loc[inv, 순매수_col] or 0) / 1e8, 1)
                                        break
                            if result["inst_buy"] is None:
                                if "기관합계" in df_sum.index:
                                    result["inst_buy"] = round(
                                        float(df_sum.loc["기관합계", 순매수_col] or 0) / 1e8, 1)
                            result["net_buy"] = round(
                                (result["foreign_buy"] or 0) + (result["inst_buy"] or 0), 1)
                except Exception:
                    pass

                # 방법A-2: 날짜별 순매수 (연속일수 계산)
                try:
                    df_daily = _s.get_market_trading_value_by_date(strt10, end, code, on="순매수")
                    if df_daily is not None and not df_daily.empty:
                        col_f = next((c for c in df_daily.columns if "외국인" in c), None)
                        col_i = next((c for c in df_daily.columns if "기관" in c), None)
                        if result["consec_days"] == 0 and (col_f or col_i):
                            consec = 0
                            for _, row in df_daily.iloc[::-1].iterrows():
                                f_val = float(row[col_f] if col_f else 0 or 0)
                                i_val = float(row[col_i] if col_i else 0 or 0)
                                if f_val > 0 or i_val > 0:
                                    consec += 1
                                else:
                                    break
                            result["consec_days"] = consec
                except Exception:
                    pass

                # 방법B 폴백: 거래량 × 현재가 근사 (방법A 완전 실패 시)
                if result["foreign_buy"] is None and price and price > 0:
                    try:
                        df_vol = _s.get_market_trading_volume_by_investor(strt5, end, code)
                        if df_vol is not None and not df_vol.empty:
                            순매수_col = next((c for c in df_vol.columns if "순매수" in c), None)
                            if 순매수_col:
                                for inv in ["외국인합계", "외국인"]:
                                    if inv in df_vol.index:
                                        result["foreign_buy"] = round(
                                            float(df_vol.loc[inv, 순매수_col] or 0) * price / 1e8, 1)
                                        break
                                for inv in ["연기금 등", "연기금등", "연기금"]:
                                    if inv in df_vol.index:
                                        result["pension_buy"] = round(
                                            float(df_vol.loc[inv, 순매수_col] or 0) * price / 1e8, 1)
                                        break
                                if "기관합계" in df_vol.index:
                                    result["inst_buy"] = round(
                                        float(df_vol.loc["기관합계", 순매수_col] or 0) * price / 1e8, 1)
                                result["net_buy"] = round(
                                    (result["foreign_buy"] or 0) + (result["inst_buy"] or 0), 1)
                    except Exception:
                        pass
            except Exception:
                pass

        # 데몬 스레드로 실행 → 10초 내 완료 안 되면 건너뜀 (프로그램 멈춤 방지)
        _t = _th.Thread(target=_do_investor_fetch, daemon=True)
        _t.start()
        _t.join(timeout=10)
        # 타임아웃 시 result 값은 None/0 유지 → 아래 네이버 폴백으로 이어짐

    # ① 폴백: 네이버 금융 investor.naver 스크래핑
    # pykrx 투자자 API 전체 실패 시 사용
    if result["foreign_buy"] is None:
        try:
            import re as _re
            _naver_url = f"https://finance.naver.com/item/investor.naver?code={code}"
            _nr = requests.get(_naver_url,
                               headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"},
                               timeout=6)
            if _nr.status_code == 200:
                from bs4 import BeautifulSoup as _BS
                _soup = _BS(_nr.text, "lxml")
                _tbl = _soup.select_one("table.type2")
                if _tbl:
                    _rows = _tbl.select("tr")
                    # 헤더 확인: 날짜 / 외국인 / 기관계 / 개인
                    _data = []
                    for _tr in _rows:
                        _tds = _tr.select("td")
                        if len(_tds) >= 4:
                            try:
                                _date = _tds[0].get_text(strip=True).replace(".", "")
                                _f    = _tds[1].get_text(strip=True).replace(",", "").replace("+", "")
                                _i    = _tds[2].get_text(strip=True).replace(",", "").replace("+", "")
                                if _re.match(r"^\d{8}$", _date) and _f not in ("", "-"):
                                    _data.append({
                                        "foreign": int(_f) if _f.lstrip("-").isdigit() else 0,
                                        "inst":    int(_i) if _i.lstrip("-").isdigit() else 0,
                                    })
                            except Exception:
                                pass
                    if _data and price and price > 0:
                        _d5 = _data[:5]
                        _f5 = sum(d["foreign"] for d in _d5)
                        _i5 = sum(d["inst"]    for d in _d5)
                        result["foreign_buy"] = round(_f5 * price / 1e8, 1)
                        result["inst_buy"]    = round(_i5 * price / 1e8, 1)
                        result["net_buy"]     = round(result["foreign_buy"] + result["inst_buy"], 1)
                        # 연속일수: 외국인 OR 기관 양수
                        _consec = 0
                        for _d in _data[:10]:
                            if _d["foreign"] > 0 or _d["inst"] > 0:
                                _consec += 1
                            else:
                                break
                        if result["consec_days"] == 0:
                            result["consec_days"] = _consec
        except Exception:
            pass

    # ② 거래량비율 + 가격 방향성 (OHLCV 25일치)
    if result["vol_ratio"] is None or result["price_up_days"] == 0:
        try:
            from pykrx import stock as _s
            df_ohl = _s.get_market_ohlcv_by_date(strt25, end, code)
            if df_ohl is not None and not df_ohl.empty:
                vol_col   = next((c for c in df_ohl.columns if "거래량" in c), None)
                close_col = next((c for c in df_ohl.columns
                                  if "종가" in c or c.lower() == "close"), None)
                open_col  = next((c for c in df_ohl.columns
                                  if "시가" in c or c.lower() == "open"), None)

                # 거래량비율
                if vol_col and result["vol_ratio"] is None:
                    vols = [v for v in df_ohl[vol_col].tolist() if v and v > 0]
                    if len(vols) >= 5:
                        v5  = sum(vols[-5:]) / 5
                        v20 = sum(vols[-20:]) / 20 if len(vols) >= 20 else sum(vols) / len(vols)
                        result["vol_ratio"] = round(v5 / v20, 2) if v20 > 0 else None

                # 가격 방향성: 최근 5일 중 상승일수 (종가 > 시가)
                if close_col and result["price_up_days"] == 0:
                    closes = df_ohl[close_col].tolist()
                    if open_col:
                        opens  = df_ohl[open_col].tolist()
                        result["price_up_days"] = sum(
                            1 for c, o in zip(closes[-5:], opens[-5:]) if c > o)
                    elif len(closes) >= 6:
                        # 전일 대비 상승 카운트
                        result["price_up_days"] = sum(
                            1 for i in range(-5, 0) if closes[i] > closes[i - 1])
        except Exception:
            pass

    # ③ FDR 폴백 — 거래량비율 (pykrx 실패 시)
    if result["vol_ratio"] is None:
        try:
            import FinanceDataReader as fdr
            from datetime import datetime as _dt
            start_dt = (_dt.now() - timedelta(days=40)).strftime("%Y-%m-%d")
            end_dt   = _dt.now().strftime("%Y-%m-%d")
            df_fdr   = fdr.DataReader(code, start_dt, end_dt)
            if df_fdr is not None and not df_fdr.empty:
                vcol = next((c for c in df_fdr.columns
                             if c.lower() in ("volume", "vol", "거래량")), None)
                if vcol:
                    vols = [v for v in df_fdr[vcol].dropna().tolist() if v > 0]
                    if len(vols) >= 5:
                        v5  = sum(vols[-5:]) / 5
                        v_b = sum(vols) / len(vols)
                        result["vol_ratio"] = round(v5 / v_b, 2) if v_b > 0 else None
        except Exception:
            pass

    # ④ 수급 등급 판정
    f_pos    = (result["foreign_buy"] or 0) > 0
    p_pos    = (result["pension_buy"] or 0) > 0
    i_pos    = (result["inst_buy"]    or 0) > 0
    v_hi     = (result["vol_ratio"]   or 0) >= 1.5
    consec   = result["consec_days"]
    up_days  = result["price_up_days"]
    short_r  = result["short_ratio"] or 0   # 공매도 잔고비중(%)

    if   f_pos and p_pos and consec >= 3 and v_hi and up_days >= 3:
        grade = "★★★"
    elif f_pos and i_pos and consec >= 2 and up_days >= 3:
        grade = "★★"
    elif (f_pos or p_pos) and up_days >= 3:
        grade = "★"
    else:
        grade = ""

    # 공매도 잔고 과다 시 등급 강등 (공매도 세력 대기 = 추가 하방 압력)
    # 5%↑: ★★★→★★ 강등 / 10%↑: ★★★·★★→★ 강등
    if short_r >= 10 and grade in ("★★★", "★★"):
        grade = "★"
    elif short_r >= 5 and grade == "★★★":
        grade = "★★"

    result["grade"]     = grade
    result["is_strong"] = grade in ("★★", "★★★")   # 예외B 기준: ★★ 이상

    # ⑤ 라벨 생성
    parts = []
    if result["foreign_buy"] is not None:
        parts.append(f"외국인{result['foreign_buy']:+.0f}억")
    if (result["pension_buy"] or 0) != 0:
        parts.append(f"연기금{result['pension_buy']:+.0f}억")
    if result["vol_ratio"] is not None:
        parts.append(f"거래량{result['vol_ratio']:.1f}배")
    if consec > 0:
        parts.append(f"연속{consec}일")
    if up_days > 0:
        parts.append(f"상승{up_days}일")
    if short_r >= 1:
        parts.append(f"공매도{short_r:.1f}%")
    result["label"] = (f"{grade} " if grade else "") + "/".join(parts)

    result["_done"] = True
    _SUPPLY_CACHE[code] = result
    return result


# ══════════════════════════════════════════════
# 3-0. 업종 분류 수집
# ══════════════════════════════════════════════

FINANCIAL_SECTORS = ["은행", "보험", "증권", "금융투자", "금융지주", "저축은행", "카드", "캐피탈"]

def _is_financial_sector(sector: str) -> bool:
    """업종명으로 금융주 여부 판단"""
    return any(kw in (sector or "") for kw in FINANCIAL_SECTORS)


def _prev_business_day() -> str:
    """
    전일 완성 영업일 날짜(YYYYMMDD) 반환.
    오늘 장중 데이터 불완전 문제 방지 — 어제 기준으로 주말 건너뜀.
    (실제 데이터 검증은 _fetch_market_df 내부에서 날짜별 재시도로 처리)
    """
    d = datetime.now() - timedelta(days=1)
    while d.weekday() >= 5:          # 토/일 건너뜀
        d -= timedelta(days=1)
    return d.strftime("%Y%m%d")


def get_sector_map() -> dict:
    """
    pykrx로 전종목 업종 분류 수집
    반환: {종목코드: 업종명}
    ② 리포트 업종 컬럼 및 ④ 금융주 감지에 활용
    캐시: 1일
    """
    cache_f = CACHE_DIR / "sector_map.json"
    today   = datetime.now().strftime("%Y-%m-%d")
    try:
        if cache_f.exists():
            c = json.loads(cache_f.read_text(encoding="utf-8"))
            if c.get("date") == today:
                return c.get("data", {})
    except Exception:
        pass

    # ── 방법 1: pykrx get_market_sector_classifications
    sector_map = {}
    try:
        from pykrx import stock as _s
        date = _prev_business_day()
        for market in ["KOSPI", "KOSDAQ"]:
            try:
                df = _s.get_market_sector_classifications(date, market=market)
                if df is None or df.empty:
                    continue
                sec_col = next((c for c in df.columns if "업종" in c), None)
                if not sec_col and len(df.columns) >= 2:
                    sec_col = df.columns[1]
                if sec_col:
                    for idx in df.index:
                        sector_map[str(idx)] = str(df.loc[idx, sec_col])
                    print(f"  [업종 pykrx] {market} {len(sector_map)}개 로드 (컬럼: {sec_col})")
                else:
                    print(f"  [업종 pykrx] {market} 업종 컬럼 없음. 컬럼: {list(df.columns)[:6]}")
            except Exception as e:
                print(f"  [업종 조회 오류-pykrx] {market}: {e}")
    except Exception as e:
        print(f"  [업종 pykrx import 오류] {e}")

    # ── 방법 2: FinanceDataReader 폴백 (pykrx 실패 시)
    if not sector_map:
        try:
            import FinanceDataReader as fdr
            for market in ["KOSPI", "KOSDAQ"]:
                try:
                    df = fdr.StockListing(market)
                    if df is None or df.empty:
                        continue
                    # 종목코드 컬럼 — 대소문자 무관 유연 매칭
                    code_col = next(
                        (c for c in df.columns
                         if c.lower() in ("symbol", "code", "종목코드")), None
                    )
                    # 업종 컬럼 — 대소문자 / 부분 매칭
                    sec_col = next(
                        (c for c in df.columns
                         if any(k in c.lower()
                                for k in ("sector", "industry", "업종", "wics"))),
                        None
                    )
                    if not code_col:
                        print(f"  [FDR-{market}] 종목코드 컬럼 없음. 실제 컬럼: {list(df.columns)[:8]}")
                        continue
                    if not sec_col:
                        print(f"  [FDR-{market}] 업종 컬럼 없음. 실제 컬럼: {list(df.columns)[:8]}")
                        continue
                    added = 0
                    for _, row in df.iterrows():
                        c = str(row[code_col]).zfill(6)
                        s = str(row[sec_col])
                        if s and s not in ("nan", "None", "", "NaN"):
                            sector_map[c] = s
                            added += 1
                    print(f"  [FDR-{market}] {added}개 업종 로드 (컬럼: {sec_col})")
                except Exception as e:
                    print(f"  [업종 조회 오류-FDR] {market}: {e}")
            if sector_map:
                print(f"  [업종 FDR 폴백 사용] 총 {len(sector_map)}개")
        except Exception as e:
            print(f"  [FDR import 오류] {e}")

    # ── 방법 3: 주요 금융주 하드코딩 (FDR도 실패 시 최소 보장)
    # 은행/증권/보험/금융지주 → PBR-ROE 모델 적용 필수
    _FINANCIAL_FALLBACK = {
        # 금융지주
        "105560": "은행", "055550": "은행", "086790": "은행",
        "316140": "은행", "138930": "은행", "139130": "은행",
        "175330": "은행", "024110": "은행", "060000": "은행",
        # 증권
        "016360": "증권", "006800": "증권", "005940": "증권",
        "008560": "증권", "039490": "증권", "003540": "증권",
        "030610": "증권", "078020": "증권", "018880": "증권",
        "001500": "증권", "001510": "증권", "003470": "증권",
        # 손해보험
        "000810": "보험", "001450": "보험", "005830": "보험",
        "000060": "보험", "002550": "보험",
        # 생명보험
        "032830": "보험", "088350": "보험",
        # 금융지주 (기타)
        "138040": "기타금융", "000370": "기타금융",
    }
    added_fb = 0
    for code, sector in _FINANCIAL_FALLBACK.items():
        if code not in sector_map:
            sector_map[code] = sector
            added_fb += 1
    if added_fb:
        print(f"  [금융주 하드코딩 보완] {added_fb}개 추가")

    # ── 업종 없으면 조용히 빈 맵 반환
    if len(sector_map) <= len(_FINANCIAL_FALLBACK):
        print(f"  [경고] 업종 분류 자동조회 실패 → 금융주 하드코딩({len(_FINANCIAL_FALLBACK)}개)으로 운영")

    try:
        cache_f.write_text(
            json.dumps({"date": today, "data": sector_map}, ensure_ascii=False),
            encoding="utf-8"
        )
    except Exception:
        pass
    print(f"  업종 분류: {len(sector_map)}개 종목")
    return sector_map


# ══════════════════════════════════════════════
# 3. pykrx 필터링
# ══════════════════════════════════════════════

def get_filtered_tickers(minority_codes: set,
                         admin_codes: set    = None,
                         critical_codes: set = None) -> list:
    """
    pykrx로 전종목 수집 후 필터링
    - 시가총액: KOSPI 500억↑, KOSDAQ 300억↑
    - 주가 1,000원↑, 거래량 > 0 (거래정지 제외)
    - 소수계좌매도 공시 종목 제외
    - 관리종목·투자경고·투자위험 종목 제외
    - 횡령·배임·불성실공시·감사거절 종목 제외 (예외 없음)
    """
    admin_codes    = admin_codes    or set()
    critical_codes = critical_codes or set()
    from pykrx import stock as _s

    # 항상 전일 완성 데이터 사용 (장중 불완전 데이터 방지)
    date = _prev_business_day()
    print(f"  기준일: {date} (전일 기준)")

    result  = []
    summary = {}

    def _fetch_market_df(market: str, date: str):
        """
        시장 데이터 수집 (3단계 폴백)
        방법1: pykrx get_market_cap_by_ticker
        방법2: 직접 KRX API 호출 (pykrx 우회)
        방법3: FinanceDataReader (자동설치)
        반환: DataFrame (컬럼: 종가, 거래량, 거래대금, 시가총액, 상장주식수, Name)
              index = 종목코드 6자리 문자열
        """
        import pandas as pd

        def _normalize(df):
            """종가·시총 컬럼 정규화 + 종가 역산"""
            # 영문 컬럼 한글 매핑
            col_map = {
                "MarketCap": "시가총액", "Mktcap": "시가총액", "marcap": "시가총액",
                "Stocks": "상장주식수", "shares": "상장주식수",
                "Volume": "거래량", "TradingVolume": "거래량",
                "TradingValue": "거래대금", "Amount": "거래대금",
                "Close": "종가", "close": "종가",
                "Name": "Name",
            }
            df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})
            # 종가 역산 (시총÷상장주식수)
            if "종가" not in df.columns or (df["종가"] == 0).all():
                if "시가총액" in df.columns and "상장주식수" in df.columns:
                    m = df["상장주식수"] > 0
                    df.loc[m, "종가"] = (
                        df.loc[m, "시가총액"] / df.loc[m, "상장주식수"]
                    ).round().astype(int)
            for c in ["종가", "거래량", "거래대금", "시가총액", "상장주식수"]:
                if c not in df.columns:
                    df[c] = 0
            return df

        # ── 방법 1: pykrx get_market_cap_by_ticker
        used_date = date
        for _ in range(5):
            try:
                df = _s.get_market_cap_by_ticker(used_date, market=market)
                if df is not None and not df.empty and len(df.columns) >= 2:
                    df = _normalize(df)
                    if "시가총액" in df.columns and df["시가총액"].sum() > 0:
                        if used_date != date:
                            print(f"(기준일→{used_date}) ", end="")
                        return df
            except Exception:
                pass
            d_b = datetime.strptime(used_date, "%Y%m%d") - timedelta(days=1)
            while d_b.weekday() >= 5:
                d_b -= timedelta(days=1)
            used_date = d_b.strftime("%Y%m%d")

        # ── 방법 2: 직접 KRX API 호출 (날짜 재시도 포함)
        print(f"(KRX직접) ", end="", flush=True)
        krx_date = date
        for _ in range(5):
            try:
                mkt_id = "STK" if market == "KOSPI" else "KSQ"
                url = "https://data.krx.co.kr/comm/bldAttendant/getJsonData.cmd"
                payload = {
                    "bld": "dbms/MDC/STAT/standard/MDCSTAT01501",
                    "mktId": mkt_id,
                    "trdDd": krx_date,
                    "share": "1",
                    "money": "1",
                    "csvxls_isNo": "false",
                }
                hdrs = {
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                    "Referer": "https://data.krx.co.kr/contents/MDC/MDI/mdidx/mdidx.jsp",
                    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
                    "Accept": "application/json, text/javascript, */*; q=0.01",
                    "X-Requested-With": "XMLHttpRequest",
                }
                resp = requests.post(url, data=payload, headers=hdrs, timeout=20)
                resp.raise_for_status()
                data_j = resp.json()
                rows = data_j.get("OutBlock_1", [])
                if rows:
                    df_krx = pd.DataFrame(rows)
                    krx_map = {
                        "ISU_SRT_CD": "code_col",
                        "ISU_ABBRV": "Name",
                        "TDD_CLSPRC": "종가",
                        "ACC_TRDVOL": "거래량",
                        "ACC_TRDVAL": "거래대금",
                        "MKTCAP": "시가총액",
                        "LIST_SHRS": "상장주식수",
                    }
                    df_krx = df_krx.rename(columns=krx_map)
                    if "code_col" in df_krx.columns:
                        df_krx = df_krx.set_index("code_col")
                    df_krx.index = df_krx.index.astype(str).str.zfill(6)
                    for c in ["종가", "거래량", "거래대금", "시가총액", "상장주식수"]:
                        if c in df_krx.columns:
                            df_krx[c] = pd.to_numeric(
                                df_krx[c].astype(str).str.replace(",", ""), errors="coerce"
                            ).fillna(0).astype(int)
                    if not df_krx.empty and "시가총액" in df_krx.columns and df_krx["시가총액"].sum() > 0:
                        if krx_date != date:
                            print(f"(기준일→{krx_date}) ", end="")
                        return df_krx
            except Exception:
                pass
            d_b = datetime.strptime(krx_date, "%Y%m%d") - timedelta(days=1)
            while d_b.weekday() >= 5:
                d_b -= timedelta(days=1)
            krx_date = d_b.strftime("%Y%m%d")
        print(f"(KRX실패) ", end="")

        # ── 방법 3: FinanceDataReader (자동설치)
        try:
            try:
                import FinanceDataReader as fdr
            except ImportError:
                print(f"(FDR설치중) ", end="", flush=True)
                import subprocess, sys
                subprocess.check_call(
                    [sys.executable, "-m", "pip", "install", "finance-datareader", "-q"],
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
                )
                import FinanceDataReader as fdr
            print(f"(FDR폴백) ", end="", flush=True)
            df_raw = fdr.StockListing(market)
            if df_raw is None or df_raw.empty:
                return None
            col_map2 = {
                "Symbol": "code_col", "Code": "code_col",
                "Close": "종가", "Volume": "거래량", "Amount": "거래대금",
                "Marcap": "시가총액", "Stocks": "상장주식수", "Name": "Name",
            }
            df_raw = df_raw.rename(columns={k: v for k, v in col_map2.items() if k in df_raw.columns})
            if "code_col" in df_raw.columns:
                df_raw = df_raw.set_index("code_col")
            df_raw.index = df_raw.index.astype(str).str.zfill(6)
            df_raw = _normalize(df_raw)
            if not df_raw.empty and df_raw["시가총액"].sum() > 0:
                return df_raw
        except Exception as e:
            print(f"(FDR실패:{e}) ", end="")

        return None

    for market in ["KOSPI", "KOSDAQ"]:
        print(f"  [{market}] 수집 중...", end=" ", flush=True)

        df_c = _fetch_market_df(market, date)

        if df_c is None or df_c.empty:
            print(f"데이터 없음 (pykrx+FDR 모두 실패), 건너뜀")
            continue

        before = len(df_c)

        # 주가 1,000원↑ 필터
        df_f = df_c[df_c["종가"] > 1000].copy()

        # 거래량 > 0 (거래정지 제외)
        if "거래량" in df_f.columns:
            df_f = df_f[df_f["거래량"] > 0]

        # 거래대금 최소 기준: 3억 이하 유동성 부족 종목 제외
        if "거래대금" in df_f.columns:
            df_f = df_f[df_f["거래대금"] >= 300_000_000]

        # 시가총액 필터
        cap_limit = 50_000_000_000 if market == "KOSPI" else 30_000_000_000
        if "시가총액" in df_f.columns:
            df_f = df_f[df_f["시가총액"] >= cap_limit]

        # 소수계좌매도 제외
        df_f = df_f[~df_f.index.isin(minority_codes)]

        # 관리종목·투자경고·투자위험 제외
        if admin_codes:
            df_f = df_f[~df_f.index.isin(admin_codes)]

        # 횡령·배임·불성실공시·감사거절 종목 제외 (예외 없음)
        if critical_codes:
            df_f = df_f[~df_f.index.isin(critical_codes)]

        after = len(df_f)
        print(f"{before}개 → {after}개 (제외: {before-after}개)")
        summary[market] = {"before": before, "after": after}

        for code in df_f.index:
            try:
                # 종목명: df_f에 Name/종목명 컬럼 있으면 우선 사용, 없으면 pykrx 조회
                if "Name" in df_f.columns and str(df_f.loc[code, "Name"]) not in ("nan", ""):
                    name = str(df_f.loc[code, "Name"])
                elif "종목명" in df_f.columns and str(df_f.loc[code, "종목명"]) not in ("nan", ""):
                    name = str(df_f.loc[code, "종목명"])
                else:
                    name = _s.get_market_ticker_name(code)
                price  = int(df_f.loc[code, "종가"])
                shares = int(df_f.loc[code, "상장주식수"]) if "상장주식수" in df_f.columns else 0
                mktcap = int(df_f.loc[code, "시가총액"]) if "시가총액" in df_f.columns else 0
                result.append({
                    "code": code, "name": name,
                    "price": price, "market": market,
                    "shares": shares, "mktcap": mktcap,
                })
            except Exception:
                pass

    total_b = sum(v["before"] for v in summary.values())
    total_a = len(result)
    print(f"  최종: {total_b}개 → {total_a}개 ({total_b-total_a}개 제외)")
    return result


# ══════════════════════════════════════════════
# 4. FnGuide XML 수집 + 재무 필터
# ══════════════════════════════════════════════

def _is_foreign_company(soup: BeautifulSoup) -> bool:
    """해외기업 국내상장 여부 판단"""
    try:
        # 자본금 통화가 원화가 아닌 경우
        fv_unit = soup.find("face_value_unit")
        if fv_unit:
            unit = fv_unit.get_text(strip=True)
            if unit and unit not in ("원", "KRW", ""):
                return True
        # 액면가 단위가 달러/위안 등인 경우
        fv = soup.find("face_value")
        if fv:
            text = fv.get_text(strip=True)
            if any(x in text for x in ["$", "USD", "CNY", "CNH", "HKD"]):
                return True
    except Exception:
        pass
    return False


def fetch_xml_with_filter(code: str) -> dict:
    """
    FnGuide XML 수집 + 재무 필터 적용
    반환: 데이터 dict (filtered=True 이면 제외 대상)
    """
    cache_f = CACHE_DIR / f"{code}.json"
    if cache_f.exists():
        age = datetime.now() - datetime.fromtimestamp(cache_f.stat().st_mtime)
        if age.days < XML_CACHE_DAYS:
            try:
                cached = json.loads(cache_f.read_text(encoding="utf-8"))
                # 캐시 유효성 체크:
                # ① ann_equity 양수값 존재 (기본 재무 데이터)
                # ② ann_dps 키 존재 (배당 데이터 — 없으면 구버전 캐시 → 재조회)
                eq  = cached.get("ann_equity", [])
                has_equity = any(v is not None and v > 0 for v in eq)
                has_dps    = "ann_dps" in cached  # 키 존재 여부만 확인 (빈 리스트도 OK)
                if has_equity and has_dps:
                    return cached
                # 구버전 캐시 → 파일 삭제 후 재조회
                try: cache_f.unlink()
                except: pass
            except Exception:
                pass

    url = f"https://comp.fnguide.com/SVO2/xml/Snapshot_all/{code}.xml"
    try:
        r    = requests.get(url, headers=HEADERS, timeout=12)
        soup = BeautifulSoup(r.content, "lxml-xml")

        def _rv(el):
            if not el: return None
            t = el.get_text(strip=True).replace(",","")
            try: return float(t)
            except: return None

        def _val(vals, idx):
            """vals 리스트에서 idx번째 <value> 텍스트를 float으로 반환"""
            if idx < len(vals):
                t = vals[idx].get_text(strip=True).replace(",", "")
                try: return float(t) if t else None
                except: return None
            return None

        # ── 해외기업 판단
        filtered  = False
        filter_reason = ""
        if _is_foreign_company(soup):
            filtered = True
            filter_reason = "해외기업"

        # ── 발행주식수 / 자기주식
        shares   = 0
        treasury = 0
        price_el = soup.find("price")
        if price_el:
            ls = price_el.find("listed_stock_1") or price_el.find("listed_stock")
            if ls:
                try: shares = float(ls.get_text(strip=True).replace(",",""))
                except: pass

        # ── 시장 구분
        market = "KOSPI"
        stxt   = soup.find("stxt_group") or soup.find("mkt_nm")
        if stxt and ("KOSDAQ" in stxt.get_text() or "코스닥" in stxt.get_text()):
            market = "KOSDAQ"

        # ── Financial Highlight 파싱
        # FnGuide XML이 named tag → positional <value> 배열 방식으로 변경됨
        # financial_highlight_ifrs_B (별도 IFRS) 컬럼 순서:
        #   [0]매출액 [1]영업이익 [2]영업이익(발표기준) [3]세전이익
        #   [4]자산총계 [5]부채총계 [6]자본총계(=지배주주지분)
        #   [7]비지배주주지분 [8]부채비율 [9]유보율
        #   [10]영업이익률 [11]순이익률 [12]ROA [13]ROE
        #   [14]EPS [15]BPS [16]DPS [17]PER [18]PBR [19]발행주식수
        IDX_OP    = 1   # 영업이익
        IDX_PRETAX= 3   # 세전이익(법인세차감전)
        IDX_EQUITY= 6   # 자본총계(지배주주지분)
        IDX_ROE   = 13  # ROE
        IDX_BPS   = 15  # BPS
        IDX_DPS   = 16  # DPS (주당배당금)

        ann_years   = []
        ann_roe     = []
        ann_equity  = []
        ann_bps     = []
        ann_dps     = []
        ann_op_prof = []
        ann_pretax  = []
        con_roe     = []

        # 실적 데이터: financial_highlight_ifrs_B 우선, D 폴백
        fg = (soup.find("financial_highlight_ifrs_B")
              or soup.find("financial_highlight_ifrs_D")
              or soup.find("financial_highlight_gongsi")
              or soup.find("financial_highlight_annual"))

        if fg:
            for rec in fg.find_all("record"):
                date_el = rec.find("date")
                if not date_el:
                    continue
                yr_txt = date_el.get_text(strip=True)
                # 미래 추정치 건너뜀 (현재연도 이후)
                try:
                    yr_int = int(yr_txt[:4])
                    if yr_int > datetime.now().year:
                        continue
                except:
                    pass

                vals = rec.find_all("value")
                # 구버전 XML: named tag 방식
                if rec.find("roe"):
                    ann_years.append(yr_txt)
                    ann_roe.append(_rv(rec.find("roe")))
                    ann_equity.append(_rv(rec.find("controlling_interest")
                                         or rec.find("equity")))
                    ann_bps.append(_rv(rec.find("bps")))
                    ann_dps.append(_rv(rec.find("dps")))
                    ann_op_prof.append(_rv(rec.find("op_profit")
                                           or rec.find("operating_profit")))
                    ann_pretax.append(_rv(rec.find("ebt")
                                          or rec.find("pretax_profit")
                                          or rec.find("income_before_tax")))
                elif vals and len(vals) >= 16:
                    # 신버전 XML: positional <value> 방식
                    roe_v = _val(vals, IDX_ROE)
                    eq_v  = _val(vals, IDX_EQUITY)
                    if roe_v is None and eq_v is None:
                        continue   # 빈 record 건너뜀
                    ann_years.append(yr_txt)
                    ann_roe.append(roe_v)
                    ann_equity.append(eq_v)
                    ann_bps.append(_val(vals, IDX_BPS))
                    ann_dps.append(_val(vals, IDX_DPS))
                    ann_op_prof.append(_val(vals, IDX_OP))
                    ann_pretax.append(_val(vals, IDX_PRETAX))

        # ── 컨센서스 ROE — consensus 섹션에서 추출
        # consensus 섹션 구조도 positional value 방식으로 변경됨
        # financial_highlight_ifrs_B 내 (E) 연도 record에서 ROE 추출
        fg2 = (soup.find("financial_highlight_ifrs_B")
               or soup.find("financial_highlight_ifrs_D"))
        if fg2:
            for rec in fg2.find_all("record"):
                date_el = rec.find("date")
                if not date_el:
                    continue
                yr_txt = date_el.get_text(strip=True)
                try:
                    yr_int = int(yr_txt[:4])
                    if yr_int <= datetime.now().year:
                        continue   # 미래 연도만
                except:
                    continue
                vals = rec.find_all("value")
                # 구버전: named tag
                if rec.find("roe"):
                    roe_el = rec.find("roe")
                    try: con_roe.append(float(roe_el.get_text(strip=True).replace(",","")))
                    except: con_roe.append(None)
                elif vals and len(vals) > IDX_ROE:
                    roe_v = _val(vals, IDX_ROE)
                    con_roe.append(roe_v)
        # consensus 섹션도 추가 확인
        cs = soup.find("consensus")
        if cs and not con_roe:
            for rec in cs.find_all("record"):
                vals = rec.find_all("value")
                if vals:
                    roe_v = _val(vals, IDX_ROE)
                    if roe_v is not None:
                        con_roe.append(roe_v)

        # ── 재무 필터 (XML 수준에서 가능한 것만)
        if not filtered:
            # 1) 영업이익 2개년 이상 연속 손실 제외
            op_vals = [v for v in ann_op_prof if v is not None]
            if len(op_vals) >= 2:
                if op_vals[-1] < 0 and op_vals[-2] < 0:
                    filtered = True
                    filter_reason = f"영업이익 2년연속손실({op_vals[-2]:.0f},{op_vals[-1]:.0f})"

        if not filtered:
            # 2) 완전 자본잠식 — 지배주주지분 ≤ 0
            eq_vals = [v for v in ann_equity if v is not None]
            if eq_vals and eq_vals[-1] is not None and eq_vals[-1] <= 0:
                filtered = True
                filter_reason = f"완전자본잠식(지배주주지분 {eq_vals[-1]:.0f}억)"

        if not filtered:
            # 3) 자본 급감 — 전년 대비 지배주주지분 50% 이상 감소
            eq_pos = [v for v in ann_equity if v is not None and v > 0]
            if len(eq_pos) >= 2 and eq_pos[-1] < eq_pos[-2] * 0.5:
                pct = (eq_pos[-1] / eq_pos[-2] - 1) * 100
                filtered = True
                filter_reason = (f"자본급감({eq_pos[-2]:.0f}→{eq_pos[-1]:.0f}억, "
                                 f"{pct:.0f}%)")

        if not filtered:
            # 5) 지배주주지분 최솟값 — 50억 미만 극소형 S-RIM 신뢰도 낮음
            eq_vals = [v for v in ann_equity if v is not None]
            if eq_vals and 0 < eq_vals[-1] < 50:
                filtered = True
                filter_reason = f"지배주주지분 소액({eq_vals[-1]:.0f}억)"

        if not filtered:
            # 4) 법인세차감전이익 손실 (최근 1년) — Stage 1 조기 필터
            pretax_vals = [v for v in ann_pretax if v is not None]
            if pretax_vals and pretax_vals[-1] < 0:
                filtered = True
                filter_reason = f"법인세차감전이익 손실({pretax_vals[-1]:.0f}억)"

        data = {
            "code":       code,
            "market":     market,
            "shares":     shares,
            "treasury":   treasury,
            "filtered":   filtered,
            "filter_reason": filter_reason,
            "ann_years":  ann_years[-5:],
            "ann_roe":    ann_roe[-5:],
            "ann_equity": ann_equity[-5:],
            "ann_bps":    ann_bps[-5:],
            "ann_dps":    ann_dps[-5:],
            "ann_op_prof":ann_op_prof[-5:],
            "ann_pretax": ann_pretax[-5:],
            "con_roe":    con_roe[:2],
            "cached_at":  datetime.now().isoformat(),
        }
        cache_f.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
        return data

    except Exception as e:
        import traceback
        err_msg = f"{type(e).__name__}: {str(e)[:120]}"
        return {"code": code, "error": err_msg, "filtered": False}


# ══════════════════════════════════════════════
# 5. 2단계 정밀 재무 필터 (SVD_Finance)
# ══════════════════════════════════════════════

def fetch_detail_filter(code: str) -> dict:
    """
    SVD_Finance에서 단기차입금, 이익잉여금, 법인세차감전이익 체크
    반환: {"pass": True/False, "reason": ...}
    """
    cache_f = CACHE_DIR / f"{code}_detail.json"
    if cache_f.exists():
        age = datetime.now() - datetime.fromtimestamp(cache_f.stat().st_mtime)
        if age.days < XML_CACHE_DAYS:
            try:
                return json.loads(cache_f.read_text(encoding="utf-8"))
            except Exception:
                pass

    try:
        url  = (f"https://comp.fnguide.com/SVO2/ASP/SVD_Finance.asp"
                f"?pGB=1&gicode=A{code}&MenuYn=Y&NewMenuID=103&stkGb=701")
        r    = requests.get(url, headers=HEADERS, timeout=12)
        soup = BeautifulSoup(r.text, "html.parser")

        def _get_row(label: str, tbl) -> list:
            """테이블에서 특정 행의 값 추출"""
            for tr in tbl.find_all("tr"):
                th = tr.find("th")
                if not th: continue
                if label in th.get_text(strip=True):
                    vals = []
                    for td in tr.find_all("td"):
                        t = td.get_text(strip=True).replace(",","")
                        try: vals.append(float(t))
                        except: vals.append(None)
                    return vals
            return []

        # 재무상태표 테이블
        bs_tbl  = soup.find("div", id="divDaechaY")
        # 손익계산서 테이블
        is_tbl  = soup.find("div", id="divSonikY")

        pass_filter = True
        reason      = ""

        if bs_tbl:
            short_borrow  = _get_row("단기차입금", bs_tbl)
            retained      = _get_row("이익잉여금", bs_tbl)

            sb = next((v for v in reversed(short_borrow) if v is not None), None)
            re = next((v for v in reversed(retained)     if v is not None), None)

            # 단기차입금 > 이익잉여금
            if sb is not None and re is not None:
                if sb > re:
                    pass_filter = False
                    reason = f"단기차입금({sb:.0f}) > 이익잉여금({re:.0f})"

        if pass_filter and is_tbl:
            pretax = _get_row("법인세차감전 계속사업이익", is_tbl)
            if not pretax:
                pretax = _get_row("법인세비용차감전순이익", is_tbl)

            # 최근 1년 이상 손실
            recent = [v for v in pretax[-3:] if v is not None]
            if recent and recent[-1] < 0:
                pass_filter = False
                reason = f"법인세차감전이익 손실({recent[-1]:.0f})"

        result = {"pass": pass_filter, "reason": reason}
        cache_f.write_text(json.dumps(result, ensure_ascii=False), encoding="utf-8")
        return result

    except Exception as e:
        return {"pass": True, "reason": f"조회실패:{e}"}  # 실패 시 통과


# ══════════════════════════════════════════════
# 6. RIM 계산
# ══════════════════════════════════════════════

def estimate_roe(ann_roe: list, con_roe: list,
                 q_roe=None,
                 strict_mode: bool = False, ke: float = 0.1026) -> float:
    """
    ROE 추정 (웹앱·스크리닝 통일 방식):
    - 컨센서스 있음: 연간3년 + 분기 + 컨센서스1년차  가중치 (1:2:3:3:3) / 12
    - 컨센서스 없음: 연간3년 + 분기                  가중치 (1:2:3:3)   /  9

    q_roe: trailing 4Q ROE (소수 or % 단위). None이면 최근연간 ROE를 proxy로 사용.
    strict_mode: 필터 조건에만 영향, 이 함수에서는 무관.
    """
    # 연간 ROE 정규화 (소수 단위)
    vals = []
    for v in (ann_roe or []):
        if v is not None:
            vals.append(v/100 if abs(v) > 2 else v)
    if not vals: return 0.0

    recent3 = vals[-3:]
    w3 = list(range(1, len(recent3)+1))  # 1, 2, 3

    # 분기 trailing ROE: 없으면 최근 연간 ROE proxy
    if q_roe is not None:
        q = q_roe/100 if abs(q_roe) > 2 else q_roe
    else:
        q = vals[-1]   # 최근 연간 ROE를 분기 proxy로 사용

    # 컨센서스 1년차
    con1 = None
    for v in (con_roe or []):
        if v:
            con1 = v/100 if abs(v) > 2 else v
            break

    if con1 is not None:
        # 컨센서스 있음: (1:2:3:3:3) / 12
        vals5 = recent3 + [q, con1]
        w5    = w3 + [3, 3]
        return sum(v*wt for v, wt in zip(vals5, w5)) / sum(w5)
    else:
        # 컨센서스 없음: (1:2:3:3) / 9
        vals4 = recent3 + [q]
        w4    = w3 + [3]
        return sum(v*wt for v, wt in zip(vals4, w4)) / sum(w4)


def is_roe_improving(ann_roe: list) -> bool:
    """ROE 개선 추세: 3년 연속 하락 or 최근 2년 음수면 False"""
    vals = [v/100 if v and v > 1 else v for v in (ann_roe or []) if v is not None]
    if len(vals) < 2: return True
    recent = vals[-3:]
    if len(recent) >= 3:
        a, b, c = recent[-1], recent[-2], recent[-3]
        if a < b < c: return False
        if a < 0 and b < 0: return False
    elif len(recent) == 2:
        a, b = recent[-1], recent[-2]
        if a < b and a < 0: return False
    return True


def calc_rim(equity, roe, ke, shares, treasury=0):
    float_shares = (shares or 0) - (treasury or 0)
    if not all([equity, ke]) or float_shares <= 0: return {}
    excess = equity * (roe - ke)
    unit   = 1e8 / float_shares
    if excess <= 0:
        fv = equity * unit
        return {"매도주가": round(fv), "적정주가": round(fv), "매수주가": round(fv), "배열": "역배열", "roe추정": round(roe*100,2)}
    val_sell = equity + excess / ke
    val_fair = equity + excess*0.9 / (1+ke-0.9)
    val_buy  = equity + excess*0.8 / (1+ke-0.8)
    return {
        "매도주가": round(val_sell*unit), "적정주가": round(val_fair*unit),
        "매수주가": round(val_buy*unit),  "배열": "정배열", "roe추정": round(roe*100,2)
    }


def calc_pbr_roe(ann_roe: list, con_roe: list, ann_bps: list, ke: float) -> dict:
    """
    금융/보험/증권 PBR-ROE 모델
    적정PBR = ROE / ke  (상한 3배)
    적정주가 = 적정PBR × BPS
    ROE: 컨센서스 1년 후 우선 → 없으면 연간 최신
    BPS: 연간 최신값 사용
    """
    def _first(lst):
        return next((v for v in (lst or []) if v is not None), None)
    def _last(lst):
        return next((v for v in reversed(lst or []) if v is not None), None)

    roe_raw = _first(con_roe) or _last(ann_roe)
    bps     = _last(ann_bps)

    if not roe_raw or not bps or bps <= 0 or ke <= 0:
        return {}

    roe_dec = roe_raw / 100 if abs(roe_raw) > 2 else roe_raw

    if roe_dec <= 0:
        return {
            "매도주가": round(bps), "적정주가": round(bps),
            "매수주가": round(bps * 0.8), "배열": "역배열",
            "roe추정": round(roe_dec * 100, 2), "모델": "PBR-ROE",
        }

    pbr  = min(roe_dec / ke, 3.0)
    fair = round(pbr * bps)
    return {
        "매도주가": round(fair * 1.2),
        "적정주가": fair,
        "매수주가": round(fair * 0.8),
        "배열": "정배열",
        "roe추정": round(roe_dec * 100, 2),
        "모델": "PBR-ROE",
    }


# ══════════════════════════════════════════════
# 7. 스팩/투자상품 종목명 필터
# ══════════════════════════════════════════════

def _is_excluded_by_name(name: str) -> bool:
    """종목명 기반 제외 (스팩·리츠·우선주·ETF·ETN·ELW·수익증권 등)"""
    n  = name.strip()
    nu = n.upper()
    # 구조적 투자상품
    if n.endswith("리츠"):                      return True
    if n.endswith("인프라펀드"):               return True
    if "스팩" in n:                             return True
    if "SPAC" in nu:                            return True
    if "맥쿼리인프라" in n:                    return True
    # 우선주 (종목명이 "우" 또는 "우B"로 끝남)
    if n.endswith("우") or n.endswith("우B"):  return True
    # ETF·ETN·ELW·파생 상품
    if "ETF" in nu:                             return True
    if "ETN" in nu:                             return True
    if "ELW" in nu:                             return True
    if "수익증권" in n:                         return True
    if "인버스" in n:                           return True
    if "레버리지" in n:                         return True
    return False


def _is_strategic_theme(name: str) -> str:
    """
    방산·원전·조선 등 정책 수혜 전략 테마 여부 확인.
    해당 테마명 반환, 미해당이면 "" 반환.
    → 엄선 모드에서 ROE 하락/컨센서스 없음으로 탈락해도 예외C로 포함
    """
    THEMES = {
        "방산": ["방산", "방위", "한화에어로", "LIG넥스", "현대로템", "빅텍",
                 "퍼스텍", "이오시스템", "한국항공우주", "한국화약"],
        "원전": ["원전", "원자력", "두산에너빌", "비에이치아이", "우진엔텍",
                 "에스엔유", "보성파워텍", "한전기술", "비앤지스틸"],
        "조선": ["조선", "HD현대중공업", "삼성중공업", "한화오션", "케이조선",
                 "현대미포조선", "HJ중공업"],
    }
    for theme, keywords in THEMES.items():
        for kw in keywords:
            if kw in name:
                return theme
    return ""


# ══════════════════════════════════════════════
# 8. 1단계 스크리닝
# ══════════════════════════════════════════════

def _build_invest_signals(ann_roe: list, con_roe: list, ann_op_prof: list,
                          ann_bps: list = None, price: float = 0) -> str:
    """
    투자 판단 보조 시그널 문자열 생성 (⑤)
    ROE 추세 / 컨센서스 여부 / 영업이익 추세 / ROE음수경고 / PBR고평가경고
    이벤트 시그널은 save_report()에서 합산
    """
    signals = []

    valid_roe = [v for v in (ann_roe or []) if v is not None]

    # ROE 음수 2년 연속 경고 (필터 아닌 주의 표시)
    if len(valid_roe) >= 2 and valid_roe[-1] < 0 and valid_roe[-2] < 0:
        signals.append("ROE음수⚠")
    elif len(valid_roe) >= 2:
        # ROE 추세
        if valid_roe[-1] > valid_roe[-2] * 1.05:
            signals.append("ROE↑")
        elif valid_roe[-1] < valid_roe[-2] * 0.95:
            signals.append("ROE↓")
        else:
            signals.append("ROE→")

    # 컨센서스 여부
    if any(v for v in (con_roe or []) if v):
        signals.append("컨센서스O")

    # 영업이익 추세
    valid_op = [v for v in (ann_op_prof or []) if v is not None]
    if len(valid_op) >= 2:
        if valid_op[-1] > 0 and valid_op[-1] > valid_op[-2] * 1.05:
            signals.append("영업이익↑")
        elif valid_op[-1] < 0:
            signals.append("영업손실⚠")

    # BPS 3년 연속 하락 경고 (자본 지속 잠식)
    if ann_bps:
        valid_bps = [v for v in ann_bps if v is not None and v > 0]
        if len(valid_bps) >= 3 and valid_bps[-1] < valid_bps[-2] < valid_bps[-3]:
            signals.append("BPS↓↓⚠")

    # PBR 고평가 경고 (필터 아닌 주의 표시)
    if ann_bps and price and price > 0:
        bps_last = next((v for v in reversed(ann_bps) if v and v > 0), None)
        if bps_last and price > bps_last * 15:
            pbr = round(price / bps_last, 1)
            signals.append(f"PBR{pbr}배↑⚠")

    return " | ".join(signals)


# ══════════════════════════════════════════════
# 8-A. 배당주 분류 / 수급강도 배치 체크
# ══════════════════════════════════════════════

def _is_dividend_stock(ann_dps_5y: list, dps_last: float, price: float,
                       min_years: int = 3, min_yield: float = 0.03):
    """
    배당주 해당 여부 판단
    Returns: (is_div: bool, div_yield_pct: float, years_with_div: int)
    - min_years: 최근 5년 중 배당 실적 최소 연수 (기본 3년)
    - min_yield: 최소 배당수익률 (기본 3%)
    """
    if not dps_last or not price or price <= 0:
        return False, 0.0, 0
    years_with_div = sum(1 for v in (ann_dps_5y or []) if v is not None and v > 0)
    div_yield = dps_last / price
    is_div = (years_with_div >= min_years) and (div_yield >= min_yield)
    return is_div, round(div_yield * 100, 1), years_with_div


def _apply_supply_scores_batch(stocks: list, label: str = "", max_workers: int = 15) -> list:
    """
    수급강도 병렬 체크 후 각 dict에 결과 추가
    추가 필드: foreign_buy, pension_buy, inst_buy, net_buy,
              vol_ratio, consec_days, price_up_days,
              grade, is_strong, 수급코멘트
    개별 스레드 타임아웃 45초 — 네트워크 지연으로 인한 무한 대기 방지
    """
    from concurrent.futures import TimeoutError as _FuturesTimeout

    if not stocks:
        return stocks
    total = len(stocks)
    print(f"  [{label}] 수급강도 체크 중 ({total}개)...", flush=True)

    def _fetch(c):
        s = get_supply_strength(c["code"], c["현재가"])
        c = dict(c)
        c["foreign_buy"]   = s.get("foreign_buy")   or 0
        c["pension_buy"]   = s.get("pension_buy")   or 0
        c["inst_buy"]      = s.get("inst_buy")      or 0
        c["net_buy"]       = s.get("net_buy")       or 0
        c["vol_ratio"]     = s.get("vol_ratio")     or 0
        c["consec_days"]   = s.get("consec_days")   or 0
        c["price_up_days"] = s.get("price_up_days") or 0
        c["수급등급"]      = s.get("grade",     "")
        c["is_strong"]     = s.get("is_strong", False)
        c["수급코멘트"]    = s.get("label",     "")
        return c

    result = []
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futs = [ex.submit(_fetch, s) for s in stocks]
        done_cnt = 0
        for fut, orig in zip(futs, stocks):
            try:
                result.append(fut.result(timeout=45))
            except _FuturesTimeout:
                # 타임아웃: 수급 데이터 없이 원본 유지
                c = dict(orig)
                c.setdefault("foreign_buy",   0)
                c.setdefault("pension_buy",   0)
                c.setdefault("inst_buy",      0)
                c.setdefault("net_buy",       0)
                c.setdefault("vol_ratio",     0)
                c.setdefault("consec_days",   0)
                c.setdefault("price_up_days", 0)
                c.setdefault("수급등급",      "")
                c.setdefault("is_strong",     False)
                c.setdefault("수급코멘트",    "")
                result.append(c)
            except Exception:
                result.append(orig)
            done_cnt += 1
            if done_cnt % 20 == 0 or done_cnt == total:
                print(f"  [{label}] 진행: {done_cnt}/{total}", flush=True)

    grade_cnt = {g: sum(1 for c in result if c.get("수급등급") == g)
                 for g in ("★★★", "★★", "★")}
    print(f"  [{label}] 완료 — "
          f"★★★:{grade_cnt['★★★']}개 ★★:{grade_cnt['★★']}개 ★:{grade_cnt['★']}개")
    return result


def run_stage1(ke: float, tickers: list, undervalue_pct: float = 0,
               strict_mode: bool = False, ev_map: dict = None,
               sector_map: dict = None,
               theme_code_names: dict = None) -> list:
    """
    1단계 Python RIM 스크리닝
    ev_map:           event_watcher.get_event_map() 결과 (종목코드 → 이벤트 목록)
    sector_map:       get_sector_map() 결과 (종목코드 → 업종명)
    theme_code_names: {종목코드: 테마명} — 네이버 동적 테마 (예외C-동적)
    반환 리스트에 예외 종목도 포함 (예외사유 필드로 구분)
    """
    ev_map           = ev_map           or {}
    sector_map       = sector_map       or {}
    theme_code_names = theme_code_names or {}
    total      = len(tickers)
    candidates = []
    filtered   = []
    done       = 0
    print(f"\n  [1단계] {total}개 병렬 스크리닝 (동시 {MAX_WORKERS}개)...")
    start = datetime.now()

    def _proc(info):
        code  = info["code"]
        name  = info["name"]
        price = info["price"]

        # ① 절대 제외 (투자상품·해외기업 — 예외 없음)
        if _is_excluded_by_name(name):
            return None, f"{name}: 투자상품 제외"

        fin = fetch_xml_with_filter(code)
        if fin.get("error"):
            return None, f"{name}: XML 오류 [{fin.get('error','')[:80]}]"
        if fin.get("filtered") and "해외" in fin.get("filter_reason", ""):
            return None, f"{name}: 해외기업"

        # ② 지배주주지분 없으면 RIM 계산 불가 → 예외도 불가
        equities = [v for v in (fin.get("ann_equity") or []) if v and v > 0]
        if not equities:
            return None, f"{name}: 지배주주지분 없음"
        equity = equities[-1]

        # ③ RIM / PBR-ROE 계산 (금융주 여부에 따라 분기)
        ann_roe    = fin.get("ann_roe", [])
        con_roe    = fin.get("con_roe", [])
        ann_bps    = fin.get("ann_bps", [])
        ann_dps    = fin.get("ann_dps", [])
        ann_op     = fin.get("ann_op_prof", [])
        shares     = info.get("shares") or fin.get("shares") or 0
        sector     = sector_map.get(code, "")
        is_fin     = _is_financial_sector(sector)

        # 예외 D/E 판단용 최신 BPS·DPS
        bps_last = next((v for v in reversed(ann_bps) if v and v > 0), 0)
        dps_last = next((v for v in reversed(ann_dps) if v and v > 0), 0)

        # 금융/보험 포함 모든 종목 동일한 S-RIM 공식 적용 (일관성)
        roe_est = estimate_roe(ann_roe, con_roe, strict_mode=strict_mode, ke=ke)
        rim     = calc_rim(equity, roe_est, ke, shares)

        if not rim or not rim.get("적정주가", 0):
            return None, f"{name}: {'PBR-ROE' if is_fin else 'RIM'} 계산 불가"

        # 투자 시그널 (⑤) — 이벤트 제외 기초 시그널
        base_signal = _build_invest_signals(ann_roe, con_roe, ann_op, ann_bps, price)

        fair  = rim.get("적정주가", 0)
        ratio = (price / fair - 1) * 100

        # ④ 후보 dict 생성 헬퍼
        def _cand(예외사유=""):
            return {
                "name": name, "code": code,
                "market":   fin.get("market", info.get("market", "")),
                "업종":     sector,
                "현재가":   price,
                "적정주가": fair,
                "매수주가": rim.get("매수주가", 0),
                "매도주가": rim.get("매도주가", 0),
                "괴리율":   round(ratio, 1),
                "roe추정":  rim.get("roe추정", 0),
                "배열":     rim.get("배열", ""),
                "계산모델": rim.get("모델", "S-RIM"),
                "stage":    1,
                "예외사유": 예외사유,
                "_base_signal": base_signal,   # 투자시그널 기초 데이터
                "dps_last":    dps_last,        # 배당 분류용
                "ann_dps_5y":  (ann_dps or [])[-5:],  # 배당 연속성 체크용
            }, None

        # ⑤ CB/BW/유상증자 확인 — ★★★ 긍정공시 없으면 제외
        dilution_evs = [e for e in ev_map.get(code, [])
                        if e.get("type") == "negative"
                        and any(kw in e.get("desc", "")
                                for kw in ("CB 발행", "BW 발행", "유상증자"))]
        if dilution_evs:
            top_pos_all = [e for e in ev_map.get(code, [])
                           if e.get("type") == "positive" and e.get("grade") == "★★★"]
            if not top_pos_all:
                kinds = "/".join(dict.fromkeys(
                    "CB" if "CB" in e.get("desc","") else
                    "BW" if "BW" in e.get("desc","") else "유상증자"
                    for e in dilution_evs
                ))
                return None, f"{name}: {kinds} 발행 (희석 위험)"

        # ⑥ 재무 필터 실패 → 예외 A/D/E/F 체크
        if fin.get("filtered"):
            filter_reason = fin.get("filter_reason", "재무필터")
            stock_events  = ev_map.get(code, [])
            pos_events    = [e for e in stock_events if e.get("type") == "positive"]

            # 예외 A-1: ★★★ 긍정공시 (자사주소각/취득/무상증자)
            top_pos = [e for e in pos_events if e.get("grade") == "★★★"]
            # 예외 A-2: ★★ 턴어라운드 신호 (잠정실적·수주·계약)
            #           영업이익 손실 필터에만 적용 (단순 일회성 반등 포착)
            turn_ev = []
            if "영업이익" in filter_reason:
                turn_ev = [e for e in pos_events
                           if e.get("grade") == "★★"
                           and any(kw in e.get("desc", "")
                                   for kw in ("실적공시", "잠정 실적", "수주", "계약 체결", "매출 증가"))]

            exc_ev = top_pos or turn_ev
            if exc_ev:  # 고평가 여부 무관하게 예외 처리
                ev_txt = f"{exc_ev[0]['grade']} {exc_ev[0]['desc']}"
                tag    = "예외A-턴어라운드" if turn_ev and not top_pos else "예외A"
                return _cand(f"{tag} [{ev_txt}] (원필터: {filter_reason})")

            # 예외 D: PBR 극저평가 (현재가 < BPS×0.5)
            if bps_last > 0 and price < bps_last * 0.5:
                pbr_v = round(price / bps_last, 2)
                return _cand(f"예외D [PBR극저평가: {pbr_v}배] (원필터: {filter_reason})")

            # 예외 E: 고배당 (배당수익률 ≥ 4%)
            if dps_last > 0 and price > 0 and dps_last / price >= 0.04:
                div_y = round(dps_last / price * 100, 1)
                return _cand(f"예외E [고배당: {div_y}%] (원필터: {filter_reason})")

            # 예외 F: 임원/대주주 장내매수 (★★ 취득 공시)
            insider_ev = [e for e in pos_events
                          if e.get("grade") == "★★"
                          and any(kw in e.get("desc", "")
                                  for kw in ("취득", "임원", "대주주", "주요주주"))]
            if insider_ev:
                ev_txt = f"{insider_ev[0]['grade']} {insider_ev[0]['desc']}"
                return _cand(f"예외F [대주주매수: {ev_txt}] (원필터: {filter_reason})")

            # 예외 C-정적: 방산·원전·조선 전략 테마 (재무필터 탈락해도 구제)
            theme = _is_strategic_theme(name)
            if theme:
                return _cand(f"예외C [전략테마: {theme}] (원필터: {filter_reason})")

            # 예외 C-동적: 네이버 동적 테마 (재무필터 탈락해도 구제)
            if code in theme_code_names:
                dyn_theme = theme_code_names[code]
                return _cand(f"예외C [동적테마: {dyn_theme}] (원필터: {filter_reason})")

            return None, f"{name}: {filter_reason}"

        # ⑦ 엄선 모드 필터 → 예외 A / B / C 체크
        strict_reason = None
        if strict_mode:
            if not any(v for v in (con_roe or [])):
                strict_reason = "컨센서스 없음"
            elif not is_roe_improving(ann_roe):
                strict_reason = "ROE 하락추세"

        if strict_reason:
            stock_events = ev_map.get(code, [])
            pos_events_s = [e for e in stock_events if e.get("type") == "positive"]

            # 예외 A: ★★★ 긍정공시 (고평가 무관)
            top_pos = [e for e in pos_events_s if e.get("grade") == "★★★"]
            if top_pos:
                ev_txt = f"{top_pos[0]['grade']} {top_pos[0]['desc']}"
                return _cand(f"예외A [{ev_txt}] (엄선: {strict_reason})")

            # 예외 C-정적: 방산·원전·조선 전략 테마 (고평가 무관)
            theme = _is_strategic_theme(name)
            if theme:
                return _cand(f"예외C [전략테마: {theme}] (엄선: {strict_reason})")

            # 예외 C-동적: 네이버 동적 테마
            if code in theme_code_names:
                dyn_theme = theme_code_names[code]
                return _cand(f"예외C [동적테마: {dyn_theme}] (엄선: {strict_reason})")

            # 예외 D: PBR 극저평가 (현재가 < BPS×0.5)
            if bps_last > 0 and price < bps_last * 0.5:
                pbr_v = round(price / bps_last, 2)
                return _cand(f"예외D [PBR극저평가: {pbr_v}배] (엄선: {strict_reason})")

            # 예외 E: 고배당 (배당수익률 ≥ 4%)
            if dps_last > 0 and price > 0 and dps_last / price >= 0.04:
                div_y = round(dps_last / price * 100, 1)
                return _cand(f"예외E [고배당: {div_y}%] (엄선: {strict_reason})")

            # 예외 F: 임원/대주주 장내매수 (★★ 취득 공시)
            insider_ev = [e for e in pos_events_s
                          if e.get("grade") == "★★"
                          and any(kw in e.get("desc", "")
                                  for kw in ("취득", "임원", "대주주", "주요주주"))]
            if insider_ev:
                ev_txt = f"{insider_ev[0]['grade']} {insider_ev[0]['desc']}"
                return _cand(f"예외F [대주주매수: {ev_txt}] (엄선: {strict_reason})")

            # 예외 B: 수급강도 (Daum API — 소수 케이스에만 호출, 고평가 무관)
            supply = get_supply_strength(code, price)
            if supply["is_strong"]:
                return _cand(f"예외B [수급강도: {supply['label']}] (엄선: {strict_reason})")

            return None, None

        # ⑧ 고평가 → 예외 체크 후 제외
        if ratio > undervalue_pct:
            stock_events = ev_map.get(code, [])
            pos_events   = [e for e in stock_events if e.get("type") == "positive"]

            # 예외 A: 긍정공시 (★★★ / ★★ / ★ 모두 포함)
            if pos_events:
                top = pos_events[0]
                ev_txt = f"{top.get('grade','')} {top.get('desc','')}"
                return _cand(f"예외A [{ev_txt}] (고평가 but 긍정공시)")

            # 예외 C-정적: 방산·원전·조선 전략 테마
            theme = _is_strategic_theme(name)
            if theme:
                return _cand(f"예외C [전략테마: {theme}] (고평가)")

            # 예외 C-동적: 네이버 동적 테마
            if code in theme_code_names:
                dyn_theme = theme_code_names[code]
                return _cand(f"예외C [동적테마: {dyn_theme}] (고평가)")

            # 예외 D: PBR 극저평가 (고평가지만 BPS 기준으로는 저평가)
            if bps_last > 0 and price < bps_last * 0.5:
                pbr_v = round(price / bps_last, 2)
                return _cand(f"예외D [PBR극저평가: {pbr_v}배]")

            # 예외 E: 고배당 (배당수익률 ≥ 4%)
            if dps_last > 0 and price > 0 and dps_last / price >= 0.04:
                div_y = round(dps_last / price * 100, 1)
                return _cand(f"예외E [고배당: {div_y}%]")

            # 예외 F: 임원/대주주 장내매수 (★★ 취득 공시) — 고평가지만 스마트머니 신호
            insider_ev_h = [e for e in pos_events
                            if e.get("grade") == "★★"
                            and any(kw in e.get("desc", "")
                                    for kw in ("취득", "임원", "대주주", "주요주주"))]
            if insider_ev_h:
                ev_txt = f"{insider_ev_h[0]['grade']} {insider_ev_h[0]['desc']}"
                return _cand(f"예외F [대주주매수: {ev_txt}] (고평가)")

            # 예외 B: 수급강도는 고평가 전종목 대상 API 호출 시 성능 부담으로
            #         엄선모드 탈락(⑦) 케이스에서만 적용
            return None, None

        # ⑨ 정상 통과
        return _cand()

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futs = {ex.submit(_proc, t): t for t in tickers}
        for fut in as_completed(futs):
            done += 1
            try:
                r, reason = fut.result()
                if r:        candidates.append(r)
                elif reason: filtered.append(reason)
            except Exception as e:
                filtered.append(str(e))
            if done % 300 == 0:
                elapsed = (datetime.now()-start).total_seconds()
                remain  = elapsed/done*(total-done)
                print(f"    {done}/{total} | 후보:{len(candidates)} | 잔여:{remain/60:.0f}분")

    elapsed = (datetime.now()-start).total_seconds()

    # 정상 종목(예외사유 없음) → 괴리율 오름차순
    # 예외 종목 → 이벤트 등급 우선, 그 다음 괴리율
    normal    = sorted([c for c in candidates if not c.get("예외사유")],
                       key=lambda x: x["괴리율"])
    exception = sorted([c for c in candidates if c.get("예외사유")],
                       key=lambda x: x["괴리율"])
    candidates = normal + exception

    n_exc = len(exception)
    print(f"\n  [1단계 완료] {total}개 → 저평가 후보 {len(normal)}개 + 예외 {n_exc}개 ({elapsed/60:.1f}분)")

    # ── 필터 이유 통계 (디버그)
    if len(candidates) == 0 and filtered:
        from collections import Counter
        reason_short = []
        for r in filtered:
            if   "XML 오류"     in r: reason_short.append("XML오류(FnGuide접속불가)")
            elif "투자상품"      in r: reason_short.append("투자상품제외")
            elif "해외기업"      in r: reason_short.append("해외기업")
            elif "지배주주지분 없음" in r: reason_short.append("지배주주지분없음")
            elif "계산 불가"     in r: reason_short.append("RIM/PBR계산불가")
            elif "영업이익"      in r: reason_short.append("영업이익2년손실")
            elif "자본잠식"      in r: reason_short.append("완전자본잠식")
            elif "자본급감"      in r: reason_short.append("자본급감")
            elif "지배주주지분 소액" in r: reason_short.append("지배주주지분소액")
            elif "법인세"        in r: reason_short.append("법인세차감전손실")
            elif "희석"          in r: reason_short.append("CB/BW/유상증자")
            else:                      reason_short.append("기타")
        cnt = Counter(reason_short).most_common(10)
        print(f"\n  [필터 통계] 총 {len(filtered)}개 제외 이유:")
        for reason, count in cnt:
            print(f"    {reason}: {count}개")
        # XML 오류 샘플 별도 출력 (실제 오류 메세지 포함)
        xml_err_samples = [r for r in filtered if "XML 오류" in r][:3]
        if xml_err_samples:
            print(f"\n  [XML 오류 샘플]")
            for r in xml_err_samples:
                print(f"    {r}")
        print(f"\n  [기타 샘플]")
        for r in filtered[:5]:
            print(f"    {r}")

    return candidates


# ══════════════════════════════════════════════
# 9. 2단계 정밀 계산 (Python S-RIM — 웹과 동일 공식)
# ══════════════════════════════════════════════

def run_stage2(candidates: list, strict_mode: bool = False) -> list:
    """
    2단계 정밀 계산 — app.py _srim_python() 완전 동일 공식 사용
    Excel/win32com 완전 제거. Excel은 다운로드용 파일 생성에만 사용(best-effort).
    """
    if not candidates: return []
    print(f"\n  [2단계] {len(candidates)}개 후보 정밀 검증 + Python S-RIM 계산...")

    sys.path.insert(0, str(BASE))
    import fnguide_collector_v4 as _col
    import app as _app
    importlib.reload(_col)
    importlib.reload(_app)

    out_dir = BASE / "OUTPUT"
    out_dir.mkdir(exist_ok=True)
    today   = datetime.now().strftime("%Y%m%d")
    results = []

    # ── ke: 루프 밖에서 1회 수집 ─────────────────────────────────────────────
    ke = _app._compute_ke({})
    print(f"  [ke] 요구수익률 {ke*100:.2f}% 확정\n")

    for i, c in enumerate(candidates, 1):
        code = c["code"]
        name = c["name"]
        print(f"  [{i}/{len(candidates)}] {name}...", end=" ", flush=True)
        try:
            # 정밀 재무 필터 (단기차입금/법인세차감전)
            df = fetch_detail_filter(code)
            if not df.get("pass", True):
                print(f"⛔ 제외: {df.get('reason','')}")
                continue

            # FnGuide 상세 수집
            data = _col.collect(name, code)
            json_path = BASE / "WORK" / f"{code}_{name}.json"
            json_path.write_text(
                json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

            # ── Python S-RIM 계산 (웹과 완전 동일) ──────────────────────────
            apt, sell, buy, meta = _app._srim_python(data, ke)

            c2 = dict(c)
            c2.update({
                "적정주가": apt,
                "매도주가": sell,
                "매수주가": buy,
                "roe추정":  round((meta.get("roe") or 0) * 100, 2),
                "할인율":   round(ke * 100, 2),
                "배열":     meta.get("배열", ""),
                "추세":     meta.get("추세", ""),
                "roe수준":  ("ROE>요구수익" if (meta.get("roe") or 0) > ke else "ROE<요구수익"),
                "stage":    2,
                "xlsx":     "",   # Excel 생성 후 덮어씀
            })
            if c2["적정주가"] and c2["현재가"]:
                c2["괴리율"] = round((c2["현재가"] / c2["적정주가"] - 1) * 100, 1)

            # ── Excel 생성 (best-effort — 다운로드용, 계산에 미사용) ─────────
            try:
                import srim_filler_v4 as _filler
                importlib.reload(_filler)
                template = BASE / "S-RIM_V33_ForwardBlock.xlsx"
                out_path = out_dir / f"{name}_SRIM_{today}.xlsx"
                _filler.fill(str(template), str(json_path), str(out_path))
                c2["xlsx"] = str(out_path)
            except Exception as xe:
                print(f"[xlsx생성실패:{xe}]", end=" ")

            results.append(c2)
            print(f"✓ 적정주가 {c2['적정주가']:,}원 (괴리율 {c2['괴리율']:+.1f}%) [Python]")

        except Exception as e:
            c["error"] = str(e)
            c["stage"] = 2
            results.append(c)
            print(f"✗ {e}")

    results.sort(key=lambda x: x.get("괴리율", 999))
    print(f"\n  [2단계 완료] {len(results)}개")
    return results


# ══════════════════════════════════════════════
# 10. 리포트 저장
# ══════════════════════════════════════════════

def save_report(results: list, path: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "저평가종목"

    headers = [
        "분류", "종목명", "코드", "시장", "업종", "현재가", "적정주가", "매수주가", "매도주가",
        "괴리율(%)", "ROE추정(%)", "배열", "추세", "계산모델",
        "예외사유", "투자시그널", "긍정이벤트", "이벤트등급",
        "수급등급", "수급코멘트", "비고",
    ]

    # 헤더 스타일
    hf  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hft = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="DDDDDD")
    bd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hf; cell.font = hft; cell.alignment = ctr; cell.border = bd
    ws.row_dimensions[1].height = 22

    # 색상 팔레트
    COLOR_GROWTH    = "E8F4EA"  # 연초록  — 성장주
    COLOR_DIVIDEND  = "F3E5F5"  # 연보라  — 배당주
    COLOR_EXCEPTION = "FFF8E1"  # 연노랑  — 예외처리
    COLOR_EVENT     = "EAF4FB"  # 연파랑  — 긍정공시 있음 (성장주에 덧씌움)
    COLOR_ERROR     = "FCE8E8"  # 연빨강  — 오류

    for row, r in enumerate(results, 2):
        stage    = r.get("stage", 1)
        예외사유 = r.get("예외사유", "")
        분류     = r.get("분류", "성장주")

        # 계산모델 표시
        if stage == 2:
            계산모델 = "V33엑셀"
        elif r.get("계산모델") == "PBR-ROE":
            계산모델 = "PBR-ROE(금융)"
        else:
            계산모델 = "Python간이"

        # 이벤트 텍스트 구성
        pos_events  = r.get("events_positive", [])
        이벤트_desc = " / ".join(
            f"[{e.get('date','')[-4:]}] {e.get('desc','')}"
            for e in pos_events
        ) if pos_events else ""
        이벤트_등급 = " / ".join(e.get("grade", "") for e in pos_events) if pos_events else ""

        # 투자시그널 = 기초시그널 + 이벤트 요약
        base_sig  = r.get("_base_signal", "")
        ev_sig    = " / ".join(
            f"{e.get('grade','')} {e.get('desc','')[:14]}"
            for e in pos_events[:3]
        ) if pos_events else ""
        투자시그널 = " | ".join(filter(None, [base_sig, ev_sig]))

        # 수급코멘트: 기본 수급코멘트 + 배당주 정보 병합
        수급코멘트 = r.get("수급코멘트", "")
        if 분류 == "배당주":
            div_info = f"배당{r.get('배당수익률_pct',0)}%/{r.get('배당연수',0)}년"
            수급코멘트 = f"{div_info}  {수급코멘트}".strip()
        elif 분류 == "성장주" and r.get("복합점수"):
            수급코멘트 = f"복합점수{r['복합점수']:.3f}  {수급코멘트}".strip()

        # 분류 표시 (배당주는 배당수익률 병기)
        분류_표시 = 분류
        if 분류 == "배당주":
            분류_표시 = f"배당주 {r.get('배당수익률_pct',0)}%"

        수급등급 = r.get("수급등급", "")

        vals = [
            분류_표시,
            r.get("name", ""),    r.get("code", ""),    r.get("market", ""),
            r.get("업종", ""),    r.get("현재가", 0),   r.get("적정주가", 0),
            r.get("매수주가", 0), r.get("매도주가", 0),
            r.get("괴리율", 0),   r.get("roe추정", 0),
            r.get("배열", ""),    r.get("추세", ""),    계산모델,
            예외사유, 투자시그널, 이벤트_desc, 이벤트_등급,
            수급등급, 수급코멘트, r.get("error", ""),
        ]

        # 행 배경색 결정 (분류 우선, 오류 최우선)
        if r.get("error"):
            bg = COLOR_ERROR
        elif 분류 == "배당주":
            bg = COLOR_DIVIDEND
        elif 분류 == "예외처리":
            bg = COLOR_EXCEPTION
        elif pos_events:
            bg = COLOR_EVENT
        else:
            bg = COLOR_GROWTH
        rf = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.fill   = rf
            cell.border = bd
            # 텍스트 컬럼 좌측 정렬
            # 1=분류,2=종목명,5=업종,15=예외사유,16=시그널,17=이벤트,19=수급등급,20=수급코멘트,21=비고
            cell.alignment = left if col in (1, 2, 5, 15, 16, 17, 19, 20, 21) else ctr
            # 괴리율 강조 (10번째)
            if col == 10 and isinstance(v, (int, float)):
                cell.font = Font(color="185FA5" if v < 0 else "C00000", bold=True)
            # 이벤트 등급 강조 (18번째)
            if col == 18 and v:
                cell.font = Font(color="7B3F00", bold=True)
            # 수급등급 강조 (19번째) ★ 별 개수에 따라 색상
            if col == 19 and v:
                clr = "8B0000" if "★★★" in str(v) else ("C0500A" if "★★" in str(v) else "8B6914")
                cell.font = Font(color=clr, bold=True)
            # 예외사유 강조 (15번째)
            if col == 15 and v:
                cell.font = Font(color="805500", bold=True, italic=True)
            # 투자시그널 강조 (16번째)
            if col == 16 and v:
                cell.font = Font(color="1A5276", size=9)
            # PBR-ROE 모델 강조 (14번째)
            if col == 14 and "PBR-ROE" in str(v):
                cell.font = Font(color="6C3483", bold=True)
            # 분류 강조 (1번째)
            if col == 1:
                if "배당주" in str(v):
                    cell.font = Font(color="6A0DAD", bold=True)
                elif v == "성장주":
                    cell.font = Font(color="1B5E20", bold=True)
                elif v == "예외처리":
                    cell.font = Font(color="7D6608", bold=True)

        # 이벤트/예외 행은 높이 추가
        ws.row_dimensions[row].height = 32 if (투자시그널 or 예외사유 or 수급코멘트) else 18

    # 열 너비 (헤더 순서와 동일)
    col_widths = [12, 16, 8, 7, 12, 10, 10, 10, 10, 10, 10, 8, 10, 12, 30, 35, 40, 12, 35, 25]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # 요약 범례
    ws.append([])
    n_growth = sum(1 for r in results if r.get("분류") == "성장주")
    n_div    = sum(1 for r in results if r.get("분류") == "배당주")
    n_exc    = sum(1 for r in results if r.get("분류") == "예외처리")
    ws.append([f"성장주 {n_growth}개  |  배당주 {n_div}개  |  예외처리 {n_exc}개  |  "
               f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])

    # 범례 색상 표시 (마지막 행 다음)
    legend_row = ws.max_row + 2
    legend = [
        (COLOR_GROWTH,    "성장주 (저평가 + 수급강도 복합점수 Top50)"),
        (COLOR_EVENT,     "성장주 + 긍정공시 있음"),
        (COLOR_DIVIDEND,  "배당주 (꾸준한 배당 + 수익률 3%↑ + 저평가)"),
        (COLOR_EXCEPTION, "예외처리 (필터탈락 but 공시/테마/PBR극저/대주주매수)"),
        (COLOR_ERROR,     "계산 오류"),
    ]
    for i, (color, label) in enumerate(legend):
        c1 = ws.cell(row=legend_row + i, column=1, value="  ")
        c1.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        c2 = ws.cell(row=legend_row + i, column=2, value=label)
        c2.font = Font(size=9, italic=True)

    wb.save(path)
    print(f"  📊 저장: {path}")


# ══════════════════════════════════════════════
# 11. 메인
# ══════════════════════════════════════════════

def run_full(undervalue_pct: float = 0, stage2: bool = True,
             strict_mode: bool = False):
    print("\n" + "="*55)
    print("  RIM 병합 스크리닝 v2")
    print(f"  저평가 기준: 현재가/적정주가 ≤ {1+undervalue_pct/100:.0%}")
    print("="*55)
    start = datetime.now()

    # pykrx 최신 버전 체크 & 자동 업그레이드
    _ensure_pykrx_latest()

    # ── 이벤트 맵 + 동적 테마 사전 로드
    sys.path.insert(0, str(BASE))
    try:
        import event_watcher as _ew
        print("\n  공시 이벤트 조회 중 (최근 10영업일)...")
        ev_map = _ew.get_event_map(days=10)
        print(f"  이벤트 보유 종목: {len(ev_map)}개")
    except Exception as _e:
        print(f"  [경고] 이벤트 조회 실패: {_e} → 예외 기능 없이 진행")
        ev_map  = {}
        _ew     = None

    # 네이버 동적 테마 (상위 5개 테마 종목 → 예외C-동적)
    theme_code_names = {}
    if _ew is not None:
        try:
            print("  동적 테마 조회 중 (네이버 상위 5개 테마)...")
            theme_code_set, themes_dict = _ew.get_dynamic_theme_code_set(top_n=5)
            # {종목코드: 테마명} 역매핑
            theme_code_names = {
                code: tname
                for tname, codes in themes_dict.items()
                for code in codes
            }
            print(f"  동적 테마 종목: {len(theme_code_names)}개 ({len(themes_dict)}개 테마)")
        except Exception as _et:
            print(f"  [경고] 동적 테마 조회 실패: {_et}")

    ke             = get_discount_rate()
    minority_codes = get_minority_sell_codes()
    print(f"\n  외국인+기관 수급 데이터 사전 수집 중 (pykrx)...")
    _load_market_supply_batch()
    print(f"  관리종목/투자경고/위험 조회 중...")
    admin_codes    = get_admin_codes()
    print(f"  횡령·배임·불성실공시·감사거절 조회 중...")
    critical_codes = get_critical_negative_codes()
    print(f"  업종 분류 조회 중...")
    sector_map     = get_sector_map()
    print(f"\n  종목 필터링 중...")
    tickers = get_filtered_tickers(
        minority_codes,
        admin_codes=admin_codes,
        critical_codes=critical_codes,
    )
    if strict_mode:
        print("  [엄선 모드] 컨센서스 있음 + ROE 개선 추세 종목만")

    candidates = run_stage1(ke, tickers, undervalue_pct, strict_mode,
                            ev_map=ev_map, sector_map=sector_map,
                            theme_code_names=theme_code_names)

    # ── 3분류: 배당주 / 성장주(수급강도 top50) / 예외처리
    normal     = [c for c in candidates if not c.get("예외사유")]
    exceptions = [c for c in candidates if c.get("예외사유")]

    # ① 배당주 분류 — 저평가 중 꾸준한 배당(최근5년 중 3년↑) + 수익률 3%↑
    dividend_stocks, growth_candidates = [], []
    for c in normal:
        is_div, div_yield, div_years = _is_dividend_stock(
            c.get("ann_dps_5y", []), c.get("dps_last", 0), c.get("현재가", 0))
        if is_div:
            c = dict(c)
            c["분류"]          = "배당주"
            c["배당수익률_pct"] = div_yield
            c["배당연수"]       = div_years
            dividend_stocks.append(c)
        else:
            growth_candidates.append(c)

    print(f"\n  [3분류] 배당주 {len(dividend_stocks)}개 / 성장주 후보 {len(growth_candidates)}개 / 예외처리 {len(exceptions)}개")

    # ② 배당주 수급강도 체크 → 수급강도 강 우선, 그 다음 배당수익률 내림차순
    if dividend_stocks:
        dividend_stocks = _apply_supply_scores_batch(dividend_stocks, "배당주")
        dividend_stocks.sort(key=lambda x: (
            -int(x.get("is_strong", False)),
            -x.get("배당수익률_pct", 0),
        ))

    # ③ 성장주: 괴리율 상위 100개 → 수급강도 체크 → 복합점수 → top 50
    pool = sorted(growth_candidates, key=lambda x: x["괴리율"])[:100]
    if pool:
        pool = _apply_supply_scores_batch(pool, "성장주")
        n = len(pool)
        # ── 5-factor 순위 정렬 (낮을수록 좋은 지표는 오름차순, 높을수록 좋은 지표는 내림차순)
        gap_order     = sorted(range(n), key=lambda i:  pool[i]["괴리율"])
        foreign_order = sorted(range(n), key=lambda i: -(pool[i].get("foreign_buy")  or 0))
        pension_order = sorted(range(n), key=lambda i: -(pool[i].get("pension_buy")  or 0))
        vol_order     = sorted(range(n), key=lambda i: -(pool[i].get("vol_ratio")    or 0))
        consec_order  = sorted(range(n), key=lambda i: -(pool[i].get("consec_days")  or 0))

        gap_rank     = {v: r for r, v in enumerate(gap_order)}
        foreign_rank = {v: r for r, v in enumerate(foreign_order)}
        pension_rank = {v: r for r, v in enumerate(pension_order)}
        vol_rank     = {v: r for r, v in enumerate(vol_order)}
        consec_rank  = {v: r for r, v in enumerate(consec_order)}
        denom = max(n - 1, 1)

        # 데이터 가용성 확인
        has_foreign = any((c.get("foreign_buy") or 0) != 0 for c in pool)
        has_pension = any((c.get("pension_buy") or 0) != 0 for c in pool)
        has_consec  = any((c.get("consec_days") or 0) != 0 for c in pool)

        if has_foreign and has_pension:
            mode = "5-factor"
            print("  [복합점수] 괴리율30%+외국인20%+연기금20%+거래량20%+연속일10%")
        elif has_foreign:
            mode = "4-factor(연기금없음)"
            print("  [복합점수] ⚠ 연기금 데이터 없음 → 괴리율30%+외국인30%+거래량25%+연속일15%")
        else:
            mode = "2-factor(수급없음)"
            print("  [복합점수] ⚠ 수급 데이터 없음 → 괴리율70%+거래량30%")

        for i, c in enumerate(pool):
            if mode == "5-factor":
                # 괴리율30% + 외국인20% + 연기금20% + 거래량20% + 연속일10%
                raw_score = (
                    gap_rank[i]     * 0.30 +
                    foreign_rank[i] * 0.20 +
                    pension_rank[i] * 0.20 +
                    vol_rank[i]     * 0.20 +
                    consec_rank[i]  * 0.10
                ) / denom
            elif mode == "4-factor(연기금없음)":
                # 괴리율30% + 외국인30% + 거래량25% + 연속일15%
                raw_score = (
                    gap_rank[i]     * 0.30 +
                    foreign_rank[i] * 0.30 +
                    vol_rank[i]     * 0.25 +
                    consec_rank[i]  * 0.15
                ) / denom
            else:
                # 괴리율70% + 거래량30%
                raw_score = (gap_rank[i] * 0.70 + vol_rank[i] * 0.30) / denom

            c["복합점수"] = round(1.0 - raw_score, 3)  # 1에 가까울수록 좋음
            c["분류"]     = "성장주"
        growth_top50 = sorted(pool, key=lambda x: -x["복합점수"])[:50]
    else:
        growth_top50 = []

    # ④ 예외처리 분류 태그
    for c in exceptions:
        if not c.get("분류"):
            c["분류"] = "예외처리"

    print(f"  [최종] 성장주 {len(growth_top50)}개 / 배당주 {len(dividend_stocks)}개 / 예외처리 {len(exceptions)}개")

    # ── Stage 2 대상: 성장주 top50 + 배당주 전체 + 예외처리 전체
    stage2_targets = growth_top50 + dividend_stocks + exceptions
    if stage2 and stage2_targets:
        print(f"  2단계 대상: {len(stage2_targets)}개")
        final = run_stage2(stage2_targets, strict_mode=strict_mode)
    else:
        final = stage2_targets

    # ── 이벤트 정보 전체 종목에 추가 (예외·정상 모두)
    if _ew and ev_map:
        try:
            final = _ew.enrich_with_events(final, ev_map)
            # enrich_with_events가 정렬을 바꾸므로 3분류 기준으로 재정렬
            성장주_f = sorted([r for r in final if r.get("분류") == "성장주"],
                              key=lambda x: -x.get("복합점수", 0))
            배당주_f = sorted([r for r in final if r.get("분류") == "배당주"],
                              key=lambda x: (-int(x.get("is_strong", False)),
                                             -x.get("배당수익률_pct", 0)))
            예외_f   = sorted([r for r in final if r.get("분류") == "예외처리"],
                              key=lambda x: (-x.get("event_score", 0), x.get("괴리율", 0)))
            기타_f   = [r for r in final if not r.get("분류")]
            final = 성장주_f + 배당주_f + 예외_f + 기타_f
        except Exception as _e2:
            print(f"  [경고] 이벤트 보강 실패: {_e2}")

    today  = datetime.now().strftime("%Y%m%d_%H%M")
    (BASE / "SCREENING").mkdir(exist_ok=True)
    report = str(BASE / "SCREENING" / f"RIM스크리닝_{today}.xlsx")
    save_report(final, report)

    elapsed = (datetime.now()-start).total_seconds()
    n_normal = sum(1 for r in final if not r.get("예외사유"))
    n_exc    = sum(1 for r in final if r.get("예외사유"))
    print(f"\n  총 소요시간: {elapsed/60:.1f}분 | 저평가: {n_normal}개 | 예외: {n_exc}개")
    return final


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser(description="RIM 스크리닝")
    p.add_argument("--pct",    type=float, default=0,   help="저평가 기준 %% (0=적정가 이하)")
    p.add_argument("--stage1",  action="store_true", help="1단계만 실행")
    p.add_argument("--strict",  action="store_true", help="엄선 모드 (컨센서스+ROE개선)")
    args = p.parse_args()
    run_full(undervalue_pct=args.pct, stage2=not args.stage1, strict_mode=args.strict)
