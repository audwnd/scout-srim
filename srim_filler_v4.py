# -*- coding: utf-8 -*-
"""
srim_filler_v4.py - collector_v4 JSON → 엑셀 매핑 (완전 수정판)

핵심 원칙:
- Data 시트 B열(행4~21) = 수식(=I열)이므로 절대 직접 쓰지 않음
- FN_Snapshot 시트에만 값 입력 → Data 수식이 자동 참조
- Data 시트 Financial Highlight(행25~) = 직접 입력 OK
"""

import argparse, json
from datetime import datetime
from typing import Optional
import openpyxl

def _f(x) -> Optional[float]:
    if x is None: return None
    try: return float(str(x).replace(",",""))
    except: return None

def _i(x) -> Optional[int]:
    v = _f(x)
    return None if v is None else int(v)

def fill(template_path, json_path, out_path, strict_mode: bool = False):
    with open(json_path, encoding="utf-8") as f:
        d = json.load(f)

    wb = openpyxl.load_workbook(template_path)
    ws_fn   = wb["FN_Snapshot"]
    ws_data = wb["Data"]

    ann = d.get("annual", {})
    con = d.get("consensus", {})
    ind = d.get("industry", {})

    act_years = ann.get("years", [])
    est_years = con.get("years", [])

    act = act_years[-5:]
    est = est_years[:3]

    def get_act(key, yr):
        yrs = ann.get("years", [])
        vals = ann.get(key, [])
        if yr in yrs:
            i = yrs.index(yr)
            return vals[i] if i < len(vals) else None
        return None

    def get_est(key, yr):
        yrs = con.get("years", [])
        vals = con.get(key, [])
        if yr in yrs:
            i = yrs.index(yr)
            return vals[i] if i < len(vals) else None
        return None

    # ─────────────────────────────────────────────
    # FN_Snapshot 시트 입력
    # Data 시트 수식이 여기서 값을 자동 참조함
    # ─────────────────────────────────────────────
    def fn(row, col, val):
        if val is not None:
            ws_fn.cell(row, col).value = val

    # 기초자료 (Data B열 수식 참조 대상)
    # 현재가: B114 (Data I5 = OFFSET(A114, 0,1) = B114)
    # Data!B5 = LEFT(I5, SEARCHB("/",I5)-1)*1 → "/" 포함 문자열 필요
    price = _i(d.get("현재가"))
    fn(26,  1, d["meta"].get("name",""))   # Data!B4 수식이 A26 참조
    fn(114, 2, f"{price:,}/ " if price else None)

    # 베타: D117 (Data I6 = OFFSET(C117, 0,1) = D117)
    fn(117, 4, _f(d.get("베타")))

    # 발행주식수: B128 (Data I7 = OFFSET(A128, 0,1) = B128)
    b = _i(d.get("발행주식수_보통", 0))
    u = _i(d.get("발행주식수_우선", 0))
    fn(128, 2, f"{b:,}/ {u:,}" if b else None)

    # 자기주식: C185 (Data I8 = OFFSET(A185, 0,2) = C185)
    fn(185, 3, _i(d.get("자기주식", 0)))

    # 가치평가 지표 (Data I열 OFFSET 참조)
    # PER: A52 (Data I16 = OFFSET(A51,1,0) = A52)
    fn(52,  1, _f(d.get("PER")))

    # 12M PER: A57 (Data I17 = OFFSET(A54,3,0) = A57)
    fn(57,  1, _f(d.get("12M_PER")))

    # 업종 PER: A64 (Data I18 = OFFSET(A59,5,0) = A64)
    fn(64,  1, _f(d.get("업종_PER")))

    # EV/EBITDA: B286 (Data I19 = OFFSET(A286,0,1) = B286)
    fn(286, 2, _f(d.get("EV_EBITDA")))

    # PBR: A71 (Data I20 = OFFSET(A70,1,0) = A71)
    fn(71,  1, _f(d.get("PBR")))

    # 배당수익률: A77 (Data I21 = OFFSET(A73,4,0) = A77)
    # FN_Snapshot은 소수 (0.0139)
    div_raw = _f(d.get("배당수익률"))
    fn(77,  1, div_raw / 100 if div_raw and div_raw > 1 else div_raw)

    # ── ROA: FN_Snapshot!B443~F443 (% 단위 그대로 입력)
    # Data!B39 수식: =OFFSET(INDIRECT(H24), 37, col-1)/100 → B443/100
    roa_list = ann.get("ROA", [])
    act_yrs_all = ann.get("years", [])
    for i, yr in enumerate(act):
        col_fn = 2 + i  # B=2 ~ F=6
        yr_i = act_yrs_all.index(yr) if yr in act_yrs_all else None
        roa_val = roa_list[yr_i] if yr_i is not None and yr_i < len(roa_list) else None
        # ROA는 % 단위로 저장됨 (예: 11.48, -8.95) → 그대로 입력 (Data 수식이 /100 처리)
        ws_fn.cell(443, col_fn).value = _f(roa_val) if _f(roa_val) is not None else ""

    # 업종비교 (행277~292)
    fn(277, 2, d["meta"].get("name",""))
    fn(282, 2, _f(ind.get("종목_PER")))
    fn(282, 3, _f(ind.get("업종_PER")))
    fn(290, 2, _f(ind.get("종목_ROE")))
    fn(290, 3, _f(ind.get("업종_ROE")))
    fn(290, 4, _f(ind.get("KOSPI_ROE")))
    fn(291, 2, _f(ind.get("종목_배당")))
    fn(291, 3, _f(ind.get("업종_배당")))
    fn(291, 4, _f(ind.get("KOSPI_배당")))

    print("  FN_Snapshot 입력 완료")

    # ─────────────────────────────────────────────
    # 결과 시트: ROE 순위 = 가중평균 고정
    # (컨센서스 대신 실적 5년 가중평균 ROE 사용)
    # ─────────────────────────────────────────────
    ws_결과 = wb["결과"]

    # ROE 추정 방식
    # - 일반 모드: 가중평균 고정 (분기 반영 자동 계산)
    # - 엄선 모드: 컨센서스 우선 (1단계 스크리닝과 일관성)
    if strict_mode:
        # 컨센서스 ROE 존재 여부 확인
        con_roe_list = con.get("ROE", [])
        con_val = next((v for v in con_roe_list if v), None)
        if con_val:
            ws_결과["C20"] = "1순위"   # 컨센서스 1년차
            print(f"  ROE 추정: 컨센서스 (엄선 모드)")
        else:
            ws_결과["C20"] = "가중평균"
            print(f"  ROE 추정: 가중평균 (컨센서스 없음)")
    else:
        ws_결과["C20"] = "가중평균"
        print(f"  ROE 추정: 가중평균 (분기 반영 시 자동 계산)")

    # ─────────────────────────────────────────────
    # 결과 시트: 할인율 (KIS BBB- 5년 수익률)
    # 결과!D14 = 요구수익률 직접입력값 (C14="적용" 상태)
    # ─────────────────────────────────────────────
    ws_결과 = wb["결과"]
    try:
        from kis_collector import get_bbb_minus_5yr
        bbb_rate = get_bbb_minus_5yr()
        ws_결과.cell(14, 4).value = bbb_rate   # 결과!D14
        print(f"  할인율(BBB- 5년): {bbb_rate*100:.2f}% → 결과!D14 입력")
    except Exception as e:
        print(f"  [경고] KIS 수집 실패: {e}")
        print(f"  → 기존 결과!D14 값 유지")

    # ─────────────────────────────────────────────
    # Data 시트 Financial Highlight (행25~)
    # 이 구간은 직접 입력 OK (수식이 아닌 입력 셀)
    # ─────────────────────────────────────────────
    def dw(row, col, val):
        # None → "" (빈칸): 엑셀에서 None쓰면 0이 되어 DIV/0! 발생
        # "" 쓰면 빈칸 처리되어 이전 종목 잔여값도 지워짐
        ws_data.cell(row, col).value = val if val is not None else ""

    # 실적 열 (B=2 ~ F=6)
    for i, yr in enumerate(act):
        col = 2 + i

        try: dw(25, col, datetime(int(yr), 12, 1))
        except: dw(25, col, yr)
        dw(26, col, int(yr))

        for key, row in [
            ("매출액",27), ("영업이익",28),
            ("당기순이익",30), ("지배주주순이익",31), ("비지배주주순이익",32),
            ("자산총계",33), ("부채총계",34), ("자본총계",35),
            ("지배주주지분",36), ("비지배주주지분",37), ("자본금",38),
        ]:
            dw(row, col, _f(get_act(key, yr)))

        dw(29, col, _f(get_act("영업이익", yr)))

        roe = get_act("ROE", yr)
        dw(40, col, _f(roe))

        dw(41, col, _f(get_act("EPS", yr)))
        dw(42, col, _f(get_act("BPS", yr)))
        dw(43, col, _f(get_act("DPS", yr)))

        div_v = get_act("배당수익률", yr)
        dw(44, col, _f(div_v))

    # F열(col=6)이 최근결산연도 - 항상 채워야 함 (RIM!B2 = Data!F26)
    if len(act) < 5:
        last_yr = act[-1]
        col = 6
        try: dw(25, col, datetime(int(last_yr), 12, 1))
        except: dw(25, col, last_yr)
        dw(26, col, int(last_yr))
        for key, row in [
            ("매출액",27), ("영업이익",28),
            ("당기순이득",30), ("지배주주순이익",31), ("비지배주주순이익",32),
            ("자산총계",33), ("부채총계",34), ("자본총계",35),
            ("지배주주지분",36), ("비지배주주지분",37), ("자본금",38),
        ]:
            dw(row, col, _f(get_act(key, last_yr)))
        dw(29, col, _f(get_act("영업이익", last_yr)))
        dw(40, col, _f(get_act("ROE", last_yr)))
        dw(41, col, _f(get_act("EPS", last_yr)))
        dw(42, col, _f(get_act("BPS", last_yr)))
        dw(43, col, _f(get_act("DPS", last_yr)))

    # 컨센서스 열 (G=7, H=8, I=9)
    for i, yr in enumerate(est[:3]):
        col = 7 + i
        try: dw(25, col, datetime(int(yr), 12, 1))
        except: dw(25, col, yr)
        dw(26, col, int(yr))

        for key, row in [
            ("매출액",27), ("영업이익",28),
            ("당기순이익",30), ("지배주주순이익",31),
            ("지배주주지분",36),
        ]:
            dw(row, col, _f(get_est(key, yr)))

        roe_e = get_est("ROE", yr)
        dw(40, col, _f(roe_e))

        dw(41, col, _f(get_est("EPS", yr)))
        dw(42, col, _f(get_est("BPS", yr)))
        dw(43, col, _f(get_est("DPS", yr)))

    # 업종비교 (행84~87)
    dw(84, 2, d["meta"].get("name",""))
    roe_s = _f(ind.get("종목_ROE"))
    roe_i = _f(ind.get("업종_ROE"))
    roe_k = _f(ind.get("KOSPI_ROE"))
    dw(85, 2, roe_s/100 if roe_s else None)
    dw(85, 3, roe_i/100 if roe_i else None)
    dw(85, 4, roe_k/100 if roe_k else None)
    dw(86, 2, _f(ind.get("종목_PER")))
    dw(86, 3, _f(ind.get("업종_PER")))
    div_s = _f(ind.get("종목_배당"))
    div_i = _f(ind.get("업종_배당"))
    div_k = _f(ind.get("KOSPI_배당"))
    dw(87, 2, div_s/100 if div_s else None)
    dw(87, 3, div_i/100 if div_i else None)
    dw(87, 4, div_k/100 if div_k else None)

    # ── BPS/ROE null 보완 (잠정실적P 미공시 → 직전 확정연도 값으로 대체)
    # 결과!G38 = Data!F42(BPS), G39 = G34/G38 → BPS=0이면 DIV/0!
    # F열이 잠정실적(P)이어서 BPS/ROE가 null인 경우 E열(직전년도)로 보완
    for row_num, key in [(42, 'BPS'), (40, 'ROE'), (36, '지배주주지분')]:
        f_val = ws_data.cell(row_num, 6).value   # F열 (최근결산)
        e_val = ws_data.cell(row_num, 5).value   # E열 (직전연도)
        if (f_val is None or f_val == "") and (e_val is not None and e_val != ""):
            ws_data.cell(row_num, 6).value = e_val
            print(f"    ※ Data!F{row_num}({key}) null → E열 값({e_val}) 보완")

    # ─────────────────────────────────────────────
    # FN_Snapshot 분기 영역 입력 (행478~)
    # Data 행50~ 수식이 FN_Snapshot에서 자동 참조
    # → Data!K61(최근4분기ROE) → 결과!H22 → 가중평균 수식 반영
    # ─────────────────────────────────────────────
    import re as _re2
    qtr = d.get("quarter", {})
    q_years = qtr.get("years", [])

    # 분기 영역 항상 초기화 (이전 종목 잔여값 제거)
    # None = 진짜 빈셀 (""는 OFFSET 수식이 0으로 읽음)
    for r_clr in [478, 482, 483, 484, 485, 486, 487, 488, 489, 490, 491, 492, 493, 510]:
        for c_clr in range(2, 7):  # B~F
            ws_fn.cell(r_clr, c_clr).value = None

    if q_years:
        FN_Q_ROWS = {
            "매출액":           482,
            "영업이익":         483,
            "영업이익(발표기준)": 484,
            "당기순이익":       485,
            "지배주주순이익":   486,
            "비지배주주순이익": 487,
            "자산총계":         488,
            "부채총계":         489,
            "자본총계":         490,
            "지배주주지분":     491,
            "비지배주주지분":   492,
            "자본금":           493,
            "ROA":              510,
        }
        # FN_Snapshot 분기 영역은 항상 B~F 5개 열을 채워야 함
        # OFFSET 수식이 빈셀(None)을 0으로 읽어서 MIN=0 → K61 DIV/0! 발생
        # → 분기 개수가 5개 미만이면 마지막 값을 F열까지 복제

        n = len(q_years)
        # 5개가 안 되면 마지막 값으로 패딩
        def _pad5(lst):
            lst = list(lst)
            while len(lst) < 5:
                lst.append(lst[-1] if lst else None)
            return lst[:5]

        q_years5 = _pad5(q_years)

        # 연도 입력 (FN_Snapshot 행478, B~F열)
        for i, yr_str in enumerate(q_years5):
            col = 2 + i
            try:
                m = _re2.search(r'(\d{4})/(\d{2})', yr_str)
                if m:
                    y, mo = int(m.group(1)), int(m.group(2))
                    ws_fn.cell(478, col).value = datetime(y, mo, 1)
                else:
                    ws_fn.cell(478, col).value = yr_str
            except:
                ws_fn.cell(478, col).value = yr_str

        # 재무 데이터 입력: 5개로 패딩 (None → 마지막 실제값 복제)
        for key, row_num in FN_Q_ROWS.items():
            vals = qtr.get(key, [])
            # None을 직전값으로 채우기 (forward fill)
            filled = []
            last_val = None
            for v in vals:
                if v is not None:
                    last_val = v
                filled.append(last_val)
            vals5 = _pad5(filled)
            for i, v in enumerate(vals5):
                col = 2 + i
                ws_fn.cell(row_num, col).value = _f(v) if v is not None else None

        print(f"  FN_Snapshot 분기 입력 완료: {q_years} (5열 패딩)")
    else:
        print("  분기 데이터 없음 → 가중평균은 연간 3개년 사용")

    print("  Data 시트 입력 완료")
    print(f"    실적연도: {act}")
    print(f"    컨센연도: {est}")
    print(f"    Data F26 (최근결산): {ws_data.cell(26,6).value}")
    print(f"    Data F40 (최근ROE):  {ws_data.cell(40,6).value}")
    print(f"    Data G40 (컨센ROE1): {ws_data.cell(40,7).value}")

    # RIM계산 시트 직접 보정
    if "RIM계산" in wb.sheetnames:
        ws_rim = wb["RIM계산"]
        last_act_yr = int(act[-1]) if act else 2025
        ws_rim.cell(7, 6).value = datetime(last_act_yr, 12, 31)
        ws_rim.cell(7, 10).value = last_act_yr + 10
        print(f"  RIM계산 F7={last_act_yr}-12-31, J7={last_act_yr+10}")

    wb.save(out_path)
    print(f"  ✓ 저장: {out_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("template")
    parser.add_argument("source")
    parser.add_argument("out")
    args = parser.parse_args()
    fill(args.template, args.source, args.out)
